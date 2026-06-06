from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User, Group
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from .forms import RegisterForm, MaterialForm, RequestForm, SystemSettingForm, NotificationSettingForm, UsedMaterialForm, LogSettingsForm, RefundableMaterialForm, RefundableMaterialUsageForm, DamageMaterialForm
from .models import Material, MaterialRequest, UserProfile, SystemSetting, NotificationSetting, UsedMaterial, MaterialMonthlyCount, InternalMessage, ActivityLog, LogSettings, MacSerialNumber, RefundableMaterial, RefundableMaterialUsage, DamageMaterial
from .utils import ensure_userprofile
from django.db.models import Sum, Q, F, Case, When, IntegerField, Count
from django.db import transaction
from django.utils import timezone
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.core.management import call_command
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
import json
import io
from io import StringIO
from django.core.paginator import Paginator
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json as _json
from django.db.models.functions import TruncDate

# Helper function to handle month-end resets
def process_month_end_reset():
    now = timezone.now()
    current_month_start = datetime(now.year, now.month, 1)
    
    # Check if this month's reset has already been processed
    system_key = f"month_reset_{now.year}_{now.month}"
    try:
        setting = SystemSetting.objects.get(key=system_key)
        return False
    except SystemSetting.DoesNotExist:
        pass
    
    # Process each material with quantity > 0 (excluding NOC materials)
    for material in Material.objects.filter(quantity__gt=0).exclude(created_by__userprofile__role='NOC'):
        # Archive the current quantity to MaterialMonthlyCount
        monthly_count, created = MaterialMonthlyCount.objects.get_or_create(
            material=material,
            month=current_month_start,
            defaults={'count': material.quantity}
        )
        
        if not created:
            monthly_count.count = material.quantity
            monthly_count.save()
        
        # Accumulate leftover quantity into Remaining_stock
        material.Remaining_stock += material.quantity
        # Reset quantity (in stock) to 0 for the new month
        material.quantity = 0
        material.save()
    
    # Mark this month's reset as processed
    SystemSetting.objects.update_or_create(
        key=system_key,
        defaults={'value': str(now), 'description': f'Month-end reset processed for {current_month_start.strftime("%B %Y")}'}
    )
    return True

def _set_jwt_cookies(response, user, tab_id):
    """Generate JWT tokens for user and attach them as HttpOnly cookies."""
    refresh = RefreshToken.for_user(user)
    access = str(refresh.access_token)
    refresh_str = str(refresh)

    jwt_cfg = getattr(settings, 'SIMPLE_JWT', {})
    secure = jwt_cfg.get('AUTH_COOKIE_SECURE', False)
    samesite = jwt_cfg.get('AUTH_COOKIE_SAMESITE', 'Lax')
    access_lifetime = jwt_cfg.get('ACCESS_TOKEN_LIFETIME').total_seconds()
    refresh_lifetime = jwt_cfg.get('REFRESH_TOKEN_LIFETIME').total_seconds()

    response.set_cookie(
        f'jwt_access_{tab_id}',
        access,
        max_age=int(access_lifetime),
        httponly=True,
        secure=secure,
        samesite=samesite,
    )
    response.set_cookie(
        f'jwt_refresh_{tab_id}',
        refresh_str,
        max_age=int(refresh_lifetime),
        httponly=True,
        secure=secure,
        samesite=samesite,
    )
    return response

def login_view(request):
    """Authenticate user and issue JWT tokens stored in HttpOnly cookies."""
    tab_id = request.GET.get('tab_id') or request.POST.get('tab_id')

    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == "POST":
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)

        if user:
            # Role check
            try:
                profile = UserProfile.objects.get(user=user)
                if profile.role not in ['Admin', 'Storekeeper', 'Branch']:
                    messages.error(request, "Access denied.")
                    return render(request, 'inventory/login.html')
            except UserProfile.DoesNotExist:
                messages.error(request, "Access denied.")
                return render(request, 'inventory/login.html')

            if not tab_id:
                import uuid
                tab_id = uuid.uuid4().hex[:8]

            request.tab_id = tab_id # Set this for middleware to catch and append to redirect
            
            # Update profile activity status
            profile.is_active = True
            profile.last_login = timezone.now()
            profile.save(update_fields=['is_active', 'last_login'])
            
            response = redirect('dashboard')
            _set_jwt_cookies(response, user, tab_id)
            return response
        else:
            messages.error(request, "Invalid credentials")

    return render(request, 'inventory/login.html')


@login_required
def logout_view(request):
    """Clear JWT cookies to log the user out."""
    response = redirect('login')
    tab_id = getattr(request, 'tab_id', None)
    
    if tab_id:
        response.delete_cookie(f'jwt_access_{tab_id}')
        response.delete_cookie(f'jwt_refresh_{tab_id}')
    else:
        # Fallback: if tab_id is missing, try to clear all JWT cookies
        for cookie_name in list(request.COOKIES.keys()):
            if cookie_name.startswith('jwt_access_') or cookie_name.startswith('jwt_refresh_'):
                response.delete_cookie(cookie_name)
    
    # Update profile activity status to False on logout
    try:
        profile = request.user.userprofile
        profile.is_active = False
        profile.save(update_fields=['is_active'])
    except Exception:
        pass
    
    # Also clear session just in case
    logout(request)
    return response


def token_refresh_view(request):
    """Silently refresh the access token using the refresh cookie.
    Called by JS before access token expires."""
    tab_id = request.GET.get('tab_id') or request.POST.get('tab_id')
    if not tab_id:
        return JsonResponse({'error': 'No tab_id provided.'}, status=400)

    refresh_token = request.COOKIES.get(f'jwt_refresh_{tab_id}')
    if not refresh_token:
        return JsonResponse({'error': 'No refresh token'}, status=401)

    try:
        refresh = RefreshToken(refresh_token)
        response = JsonResponse({'status': 'ok'})
        
        from django.conf import settings
        jwt_cfg = getattr(settings, 'SIMPLE_JWT', {})
        secure = jwt_cfg.get('AUTH_COOKIE_SECURE', False)
        samesite = jwt_cfg.get('AUTH_COOKIE_SAMESITE', 'Lax')
        access_lifetime = jwt_cfg.get('ACCESS_TOKEN_LIFETIME').total_seconds()
        
        response.set_cookie(
            f'jwt_access_{tab_id}',
            str(refresh.access_token),
            max_age=int(access_lifetime),
            httponly=True,
            secure=secure,
            samesite=samesite,
        )
        return response
    except Exception as e:
        # Refresh token is also invalid — force re-login
        response = JsonResponse({'error': 'Session expired. Please login again.'}, status=401)
        response.delete_cookie(f'jwt_access_{tab_id}')
        response.delete_cookie(f'jwt_refresh_{tab_id}')
        return response

@login_required
def dashboard(request):
    process_month_end_reset()
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return redirect('noc:dashboard')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'toggle_status' and role == 'Admin':
            user_id = request.POST.get('user_id')
            if user_id:
                try:
                    toggle_user = User.objects.get(id=user_id)
                    if toggle_user.is_superuser:
                        messages.error(request, "Cannot deactivate superuser accounts.")
                    else:
                        toggle_user.is_active = not toggle_user.is_active
                        toggle_user.save()
                        status_text = "activated" if toggle_user.is_active else "deactivated"
                        messages.success(request, f"User '{toggle_user.username}' {status_text}.")
                except User.DoesNotExist:
                    messages.error(request, "User not found.")
            return redirect('dashboard')

    # Request send by Branch materials approved by admin and auto update total materials count unique materials False
    now = timezone.now()
    if role == 'Branch':
        # For Branch: total_materials will be recalculated after available_quantity is computed
        # (see below after technician_approved_materials is built)
        total_materials = 0  # placeholder, updated later
    else:
        # For Admin & Storekeeper: Total count of all materials in system
        if role == 'Storekeeper':
            total_materials = Material.objects.filter(quantity__gt=0).count()  # In stock materials for Storekeeper
        else:
            total_materials = Material.objects.count()
    
    # active_tasks = Task.objects.filter(status='In Progress').count()
    if role in ['Admin', 'Storekeeper']:
        pending_requests_qs = MaterialRequest.objects.filter(
            status='Pending',
            is_archived=False,
            requested_at__year=now.year,
            requested_at__month=now.month,
        )
    else:
        pending_requests_qs = MaterialRequest.objects.filter(
            status='Pending',
            requester=request.user,
            is_archived=False,
            requested_at__year=now.year,
            requested_at__month=now.month,
        )
    
    pending_requests = pending_requests_qs.count()

    # Data for dashboard modals - Role-specific
    # all_tasks = Task.objects.all().order_by('-created_at')
    all_requests = MaterialRequest.objects.filter(requester=request.user, is_archived=False).order_by('-requested_at')
    all_used_materials = UsedMaterial.objects.filter(is_archived=False).select_related('technician', 'material').order_by('-added_at')[:10]  # Limit to 10 most recent
    
    # Role-specific material data for the materials modal
    technician_approved_materials = None
    advance_materials = None
    all_materials = None
    my_stock_count = 0
    used_materials_count = 0
    used_materials_counts = 0
    used_material_form = None
    total_price1 = 0
    
    if role == 'Branch':
        # For Branch: show only materials that completed full workflow
        # 1. Admin approved and storekeeper passed on, then branch received
        # 2. NOC-approved requests bypass storekeeper pass_on and can be received directly by branch
        # NOTE: NOT affected by monthly reset - materials persist across all months
        approved_qs = MaterialRequest.objects.filter(
            requester=request.user,
            status='Received',  # Only received status (workflow complete)
            received_by__isnull=False,  # Branch must have received it
            requested_at__year=now.year,
            requested_at__month=now.month
        ).filter(
            Q(pass_on__isnull=False) |
            Q(material__created_by__userprofile__role='NOC')
        ).select_related('material').order_by('requested_at') # Order by oldest first for FIFO consumption

        # For each material, get the total used/refundable/damaged amount for the current month
        used_totals = {}
        used_qs = UsedMaterial.objects.filter(
            technician=request.user,
            status='Accepted'
        ).values('material_id').annotate(total=Sum('quantity'))
        for u in used_qs:
            used_totals[u['material_id']] = u['total'] or 0

        # RefundableMaterial records are stored as free-text material names and are not linked to Material objects.
        # Therefore this dashboard stock calculation skips direct RefundableMaterial aggregation by material_id.

        # Include DamageMaterial (Pending/Confirmed) totals for this branch
        dam_qs = DamageMaterial.objects.filter(
            branch_user=request.user,
            status__in=['Pending', 'Confirmed']
        ).values('material_id').annotate(total=Sum('quantity'))
        for d in dam_qs:
            used_totals[d['material_id']] = used_totals.get(d['material_id'], 0) + (d['total'] or 0)

        # Get all active serials for this user to display in the modal
        all_user_serials = MacSerialNumber.objects.filter(
            assigned_to=request.user,
            status='Active'
        )
        serials_by_material = {}
        for s in all_user_serials:
            if s.material_id not in serials_by_material:
                serials_by_material[s.material_id] = []
            serials_by_material[s.material_id].append({'mac_serial': s.mac_serial, 'id': s.id})

        # Process requests and handle serialized materials
        technician_approved_materials = []
        for req in approved_qs:
            mat_id = req.material.id
            available_for_this_req = req.quantity
            
            # Deduct used quantity (FIFO)
            if mat_id in used_totals and used_totals[mat_id] > 0:
                amount_to_deduct = min(used_totals[mat_id], req.quantity)
                available_for_this_req -= amount_to_deduct
                used_totals[mat_id] -= amount_to_deduct
            
            # Get serials for this specific material
            serials = serials_by_material.get(mat_id, [])
            
            if serials:
                # Split into individual rows for each serial
                from types import SimpleNamespace
                for serial_obj in serials:
                    # Create a "virtual" request object for display
                    virtual_req = SimpleNamespace(
                        id=req.id,
                        material=req.material,
                        requested_at=req.requested_at,
                        available_quantity=1,
                        serials_display=serial_obj['mac_serial'],
                        serials_display_id=serial_obj['id'],
                        is_serialized=True
                    )
                    technician_approved_materials.append(virtual_req)
            else:
                # If no serials, show as a single aggregate row
                req.available_quantity = available_for_this_req
                req.serials_display = "N/A"
                req.is_serialized = False
                technician_approved_materials.append(req)
        
        # Sort by date (newest first)
        technician_approved_materials.reverse()

        # Get current-month Advance requests for branch user.
        # Status is kept visible so approved/rejected rows remain stable in the modal.
        advance_materials = MaterialRequest.objects.filter(
            requester=request.user,
            request_type='Advance',
            is_archived=False,
            requested_at__year=now.year,
            requested_at__month=now.month,
        ).select_related('material').order_by('-requested_at')
    else:
        # For Admin & Storekeeper: Get all materials
        if role == 'Storekeeper':
            # Storekeeper sees only in-stock materials in the main list
            all_materials = Material.objects.filter(quantity__gt=0).order_by('-added_at')
        else:
            all_materials = Material.objects.all().order_by('-added_at')
        # Total accepted used materials count
        used_materials_count = UsedMaterial.objects.filter(status='Accepted', is_archived=False).count()
        # Get all advance requests
        advance_materials = MaterialRequest.objects.filter(
            request_type='Advance',
            is_archived=False,
            requested_at__year=now.year,
            requested_at__month=now.month,
        ).select_related('material', 'requester').order_by('-requested_at')
    
    # Branch specific stats
    if role == 'Branch':
        now = timezone.now()
        # Calculate stock: Approved Requests (In) - Used Materials (Out) for current month
        total_in = MaterialRequest.objects.filter(
            requester=request.user, 
            status='Received',
            is_archived=False
        ).aggregate(s=Sum('quantity'))['s'] or 0
        
        total_out = UsedMaterial.objects.filter(
            technician=request.user,
            is_archived=False
        ).aggregate(s=Sum('quantity'))['s'] or 0
        
        my_stock_count = total_in - total_out
        
        # Used materials quantity sum for current month
        used_materials_count = UsedMaterial.objects.filter(
            technician=request.user,
            status='Accepted',
            is_archived=False
        ).aggregate(s=Sum('quantity'))['s'] or 0
        # Used materials record count for current month (number of entries)
        used_materials_count = UsedMaterial.objects.filter(
            technician=request.user,
            status='Accepted',
            is_archived=False
        ).count()
        used_material_form = UsedMaterialForm(user=request.user)

    # Total users - visible to all roles on dashboard
    all_users_list = User.objects.all().select_related('userprofile')
    total_users = all_users_list.count()
    
    #Materials monitoring show Branch user used materials count

    # Calculate low stock materials
    low_stock_materials = 0
    low_stock_material_list = []

    if role == 'Branch':
        # For Branch: split technician_approved_materials into in-stock and out-of-stock.
        # - available_quantity > 0  → stays in technician_approved_materials (shown in table)
        # - available_quantity == 0 → moved to low_stock_material_list (hidden from table)
        # total_materials reflects only the in-stock count.
        if technician_approved_materials:
            in_stock_list = []
            for req in technician_approved_materials:
                if req.available_quantity == 0:
                    low_stock_materials += 1
                    low_stock_material_list.append(req)
                else:
                    in_stock_list.append(req)
            technician_approved_materials = in_stock_list
            total_materials = len(in_stock_list)
    else:
        # For Admin/Storekeeper: Materials with status 'Low Stock' or 'Out of Stock'
        if role == 'Storekeeper':
            # For Storekeeper: Only 'Out of Stock' materials are shown as low stock
            low_stock_items = Material.objects.filter(status='Out of Stock')
        else:
            low_stock_items = Material.objects.filter(Q(status='Low Stock') | Q(status='Out of Stock'))
        low_stock_materials = low_stock_items.count()
        low_stock_material_list = low_stock_items

    # Materials monitoring for Admin
    materials_monitoring = []
    if role == 'Admin':
        branch_users = User.objects.filter(userprofile__role='Branch')
        for branch_user in branch_users:
            # Only monitor materials that completed full workflow
            # NOTE: NOT affected by monthly reset - materials persist across all months
            approved_qs = MaterialRequest.objects.filter(
                requester=branch_user,
                status='Received',
                received_by__isnull=False,  # Branch must have received it
                requested_at__year=now.year,
                requested_at__month=now.month
            ).filter(
                Q(pass_on__isnull=False) |
                Q(material__created_by__userprofile__role='NOC')
            ).select_related('material').order_by('requested_at')
            
            used_totals = {}
            used_qs = UsedMaterial.objects.filter(
                technician=branch_user,
                status='Accepted',
                added_at__year=now.year,
                added_at__month=now.month,
                is_archived=False
            ).values('material_id').annotate(total=Sum('quantity'))
            for u in used_qs:
                used_totals[u['material_id']] = u['total'] or 0

            # Include refundable totals for this branch
            # RefundableMaterial records are free-text and not linked to Material objects,
            # so skip this aggregation here to avoid invalid queries.

            # Include damaged totals for this branch
            dam_qs = DamageMaterial.objects.filter(
                branch_user=branch_user,
                status__in=['Pending', 'Confirmed']
            ).values('material_id').annotate(total=Sum('quantity'))
            for d in dam_qs:
                used_totals[d['material_id']] = used_totals.get(d['material_id'], 0) + (d['total'] or 0)
            
            for req in approved_qs:
                mat_id = req.material.id
                available_for_this_req = req.quantity
                
                if mat_id in used_totals and used_totals[mat_id] > 0:
                    amount_to_deduct = min(used_totals[mat_id], req.quantity)
                    available_for_this_req -= amount_to_deduct
                    used_totals[mat_id] -= amount_to_deduct
                
                if available_for_this_req > 0 and req.material.status == 'Normal':
                    materials_monitoring.append({
                        'branch': branch_user,
                        'material': req.material,
                        'quantity': available_for_this_req,
                        'status': req.material.status,
                        'date': req.requested_at,
                    })

    # Common stats for all roles: Unread Messages and Report Summaries
    unread_messages_count = InternalMessage.objects.filter(receiver=request.user, is_read=False).count()
    
    month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    if role == 'Branch':
        total_qty_issued = MaterialRequest.objects.filter(
            requester=request.user, status='Received', requested_at__gte=month_start
        ).aggregate(total=Sum('quantity'))['total'] or 0
        advance_count = MaterialRequest.objects.filter(
            requester=request.user, request_type='Advance', requested_at__gte=month_start
        ).count()
        total_used_qty = UsedMaterial.objects.filter(
            technician=request.user, status='Accepted', added_at__gte=month_start
        ).aggregate(total=Sum('quantity'))['total'] or 0
    else:
        # Admin/Storekeeper
        total_qty_issued = MaterialRequest.objects.filter(
            status='Received', requested_at__gte=month_start
        ).aggregate(total=Sum('quantity'))['total'] or 0
        advance_count = MaterialRequest.objects.filter(
            request_type='Advance',
            is_archived=False,
            requested_at__gte=month_start
        ).count()
        total_used_qty = UsedMaterial.objects.filter(
            status='Accepted', added_at__gte=month_start
        ).aggregate(total=Sum('quantity'))['total'] or 0
        
        #Total price
        total_price_agg1 = Material.objects.aggregate(total=Sum('total_price'))['total']
        total_price1 = total_price_agg1 if total_price_agg1 is not None else 0

    # Query Refundable and Damaged materials for the Destroy Materials Modal
    if role in ['Admin', 'Storekeeper']:
        refundable_materials = RefundableMaterial.objects.all().select_related('branch_user').order_by('-added_at')[:10]
        damaged_materials = DamageMaterial.objects.all().select_related('branch_user', 'material', 'confirmed_by').order_by('-added_at')[:10]
        branch_users = User.objects.select_related('userprofile').filter(userprofile__role='Branch').order_by('username')
    else:
        refundable_materials = RefundableMaterial.objects.filter(branch_user=request.user).select_related('branch_user').order_by('-added_at')[:10]
        damaged_materials = DamageMaterial.objects.filter(branch_user=request.user).select_related('material').order_by('-added_at')[:10]
        branch_users = None

    refundable_form = RefundableMaterialForm(user=request.user)
    damaged_form = DamageMaterialForm(user=request.user)

    return render(request, 'inventory/dashboard.html', {
        'total_materials': total_materials,
        # 'active_tasks': active_tasks,
        'pending_requests': pending_requests,
        'all_materials': all_materials,
        'technician_approved_materials': technician_approved_materials,
        'advance_materials': advance_materials,
        # 'all_tasks': all_tasks,
        'all_requests': all_requests,
        'all_used_materials': all_used_materials,
        'role': role,
        'user': request.user,
        'my_stock_count': my_stock_count,
        'used_materials_count': used_materials_count,
        'used_material_form': used_material_form,
        'total_users': total_users,
        'all_users_list': all_users_list,
        'low_stock_materials': low_stock_materials,
        'low_stock_material_list': low_stock_material_list,
        'pending_requests_list': pending_requests_qs.select_related('requester', 'material').order_by('-requested_at').filter(is_archived=False),
        'materials_monitoring': materials_monitoring,
        'unread_messages_count': unread_messages_count,
        'total_qty_issued': total_qty_issued,
        'advance_count': advance_count,
        'total_used_qty': total_used_qty,
        'total_price1': total_price1,
        'refundable_materials': refundable_materials,
        'damaged_materials': damaged_materials,
        'refundable_form': refundable_form,
        'damaged_form': damaged_form,
        'branch_users': branch_users,
    })


@login_required
def materials_monitoring_view(request):
    """Real-time materials monitoring for Admin: branch users and used materials (Django Channels)."""
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else None
    
    if role == 'NOC':
        return redirect('noc:dashboard')
    if role != 'Admin':
        messages.error(request, 'Only Admin can access Materials Monitoring.')
        return redirect('dashboard')
    
    ws_scheme = 'wss' if request.scheme == 'https' else 'ws'
    ws_host = request.get_host()
    ws_path = '/ws/inventory/materials-monitoring/'
    ws_url = f'{ws_scheme}://{ws_host}{ws_path}'
    
    # Get branch users' approved materials that haven't been fully used
    materials_monitoring = []
    branch_users = User.objects.filter(userprofile__role='Branch')
    branch_list = branch_users
    now = timezone.now()
    for branch_user in branch_users:
        # Only show materials that completed full workflow
        # NOTE: NOT affected by monthly reset - materials persist across all months
        approved_qs = MaterialRequest.objects.filter(
            requester=branch_user,
            status='Received',
            received_by__isnull=False,  # Branch must have received it
            requested_at__year=now.year,
            requested_at__month=now.month
        ).filter(
            Q(pass_on__isnull=False) |
            Q(material__created_by__userprofile__role='NOC')
        ).select_related('material').order_by('requested_at')

        used_totals = {}
        used_qs = UsedMaterial.objects.filter(
            technician=branch_user,
            status='Accepted',
            added_at__year=now.year,
            added_at__month=now.month,
            is_archived=False
        ).values('material_id').annotate(total=Sum('quantity'))
        for u in used_qs:
            used_totals[u['material_id']] = u['total'] or 0

        # RefundableMaterial records are free-text material names and are not linked to Material objects.
        # Therefore this dashboard stock calculation skips direct RefundableMaterial aggregation by material_id.

        # Include damaged totals for this branch
        dam_qs = DamageMaterial.objects.filter(
            branch_user=branch_user,
            status__in=['Pending', 'Confirmed']
        ).values('material_id').annotate(total=Sum('quantity'))
        for d in dam_qs:
            used_totals[d['material_id']] = used_totals.get(d['material_id'], 0) + (d['total'] or 0)
        
        for req in approved_qs:
            mat_id = req.material.id
            available_for_this_req = req.quantity
            
            if mat_id in used_totals and used_totals[mat_id] > 0:
                amount_to_deduct = min(used_totals[mat_id], req.quantity)
                available_for_this_req -= amount_to_deduct
                used_totals[mat_id] -= amount_to_deduct
            
            if available_for_this_req > 0 and req.material.status == 'Normal':
                materials_monitoring.append({
                    'branch': {
                        'id': branch_user.id,
                        'username': branch_user.username,
                        'full_name': branch_user.get_full_name(),
                    },
                    'material': {
                        'id': req.material.id,
                        'name': req.material.name,
                    },
                    'quantity': available_for_this_req,
                    'status': req.material.status,
                    'date': req.requested_at.isoformat(),
                })
    
    return render(request, 'inventory/materials_monitoring.html', {
        'role': role,
        'ws_url': ws_url,
        'materials_monitoring': materials_monitoring,
        'branch_list': branch_list,
    })


@login_required
def materials_view(request):
    """Materials management: Admin and Storekeeper can create, edit, delete; others read-only."""
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return redirect('noc:dashboard')

    # Base queryset - Admin/Storekeeper see all; Branch sees all (read-only)
    materials = Material.objects.all().order_by('-added_at')

    # Stock counts (Admin/Storekeeper only)
    total_normal_stock = Material.objects.filter(status='Normal').count()
    total_low_stock = Material.objects.filter(status='Low Stock').count()
    total_out_of_stock = Material.objects.filter(status='Out of Stock').count()

    # Search: name, category, status
    search = request.GET.get('search', '').strip()
    if search:
        materials = materials.filter(
            Q(name__icontains=search) | Q(category__icontains=search) | Q(status__icontains=search)
        )

    # Filter by category and stock_status
    category = request.GET.get('category', '')
    stock_status = request.GET.get('stock_status', '')
    if category:
        materials = materials.filter(category=category)
    if stock_status:
        status_map = {'low': 'Low Stock', 'normal': 'Normal', 'out_of_stock': 'Out of Stock'}
        db_status = status_map.get(stock_status, stock_status)
        materials = materials.filter(status=db_status)

    # Pagination
    paginator = Paginator(materials, 20)
    page_number = request.GET.get('page')
    materials_page = paginator.get_page(page_number)

    # POST: Create, Edit, Delete (Admin/Storekeeper only)
    if request.method == 'POST':
        if role != 'Storekeeper':
            messages.error(request, "Only Storekeeper can add, edit, or delete materials.")
            return redirect('materials')

        action = request.POST.get('action')
        material_id = request.POST.get('material_id', '').strip()

        # Delete Material (Storekeeper can delete any material)
        # if action == 'delete':
        #     if not material_id or not material_id.isdigit():
        #         messages.error(request, "Invalid material specified.")
        #         return redirect('materials')
        #     try:
        #         mat = Material.objects.get(pk=material_id)
        #         mat.soft_delete()
        #         messages.success(request, "Material deleted successfully.")
        #     except Material.DoesNotExist:
        #         messages.error(request, "Material not found.")
        #     return redirect('materials')

        # Create/Edit Material (Storekeeper can edit any material)
        instance = None
        if material_id and material_id != 'undefined' and material_id.isdigit():
            try:
                instance = Material.objects.get(pk=material_id)
                # Check if material was created by NOC - Storekeeper cannot edit NOC materials
                if instance.created_by and hasattr(instance.created_by, 'userprofile') and instance.created_by.userprofile.role == 'NOC':
                    messages.error(request, "Restricted: Materials created by NOC cannot be edited by Storekeeper.")
                    return redirect('materials')
            except Material.DoesNotExist:
                messages.error(request, "Material not found.")
                return redirect('materials')

        form = MaterialForm(request.POST, user=request.user, instance=instance)
        if form.is_valid():
            material = form.save(commit=False)
            is_new = not material.id

            material.save()
            messages.success(request, "Material saved successfully!")
            return redirect('materials')
        else:
            error_msg = "Please correct the errors below." if form.errors else "Invalid data."
            for field, errors in form.errors.items():
                if errors:
                    messages.error(request, f"{field}: {errors[0]}")
                    break
            else:
                messages.error(request, error_msg)
            return redirect('materials')

    form = MaterialForm(user=request.user)
    type_summary = []
    if role == 'Storekeeper':
        type_summary = list(Material.objects.values('Type').annotate(
            in_stock=Sum('quantity'),
            remaining_stock=Sum('Remaining_stock'),
            min_stock=Sum('min_stock_level')
        ))

    #Total price
    total_price_agg = materials.aggregate(total=Sum('total_price'))['total']
    total_price = total_price_agg if total_price_agg is not None else 0
    
    context = {
        'search': search,
        'category': category,
        'stock_status': stock_status,
        'total_normal_stock': total_normal_stock,
        'total_low_stock': total_low_stock,
        'total_out_of_stock': total_out_of_stock,
        'materials': materials_page,
        'form': form,
        'role': role,
        'user': request.user,
        'materials_page': materials_page,
        'type_summary': type_summary,
        'total_price': total_price,
    }
    return render(request, 'inventory/materials.html', context)


def materials_export_excel(request):
    """Export all materials data to Excel with latest monthly updates."""
    # Check if user is authenticated via JWT
    if not hasattr(request, 'user') or not request.user.is_authenticated:
        return HttpResponse('Authentication required.', status=401)
    
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role not in ['Admin', 'Storekeeper']:
        return HttpResponse('Permission denied. Only Admin and Storekeeper can export materials data.', status=403)

    # Get all materials with latest data
    materials = Material.objects.all().select_related('created_by').order_by('category', 'name')

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ISP-Inventory Materials Inventory'

    # Style definitions
    h_fill = PatternFill('solid', fgColor='4F46E5')
    h_font = Font(color='FFFFFF', bold=True, size=11)
    h_align = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f'Materials Inventory Report - {timezone.now().strftime("%B %Y")}'
    title_cell.font = Font(bold=True, size=14, color='1E1B4B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ['ID', 'Material Name', 'Category', 'Type', 'In Stock','Rate', 'Total Price', 'Remaining Stock', 'Min Stock Level', 'Status', 'Created By', 'Last Updated']
    ws.append([])
    ws.append(headers)

    # Style header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Adjust specific column widths
    ws.column_dimensions['B'].width = 25  # Material Name
    ws.column_dimensions['C'].width = 15  # Category
    ws.column_dimensions['I'].width = 20  # Created By
    ws.column_dimensions['J'].width = 18  # Last Updated

    # Data rows
    for material in materials:
        # Get creator info with role
        if material.created_by:
            try:
                role = material.created_by.userprofile.role
                creator_info = f"{material.created_by.username} [{role}]"
            except:
                creator_info = material.created_by.username
        else:
            creator_info = 'Storekeeper'  # Default to Storekeeper if created_by is null
        
        row_data = [
            material.id,
            material.name,
            material.category,
            material.Type,
            material.quantity,
            material.rate,
            material.total_price,
            material.Remaining_stock,
            material.min_stock_level,
            material.status,
            creator_info,
            material.updated_at.strftime('%Y-%m-%d'),
        ]
        ws.append(row_data)

        # Style data row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = thin
            cell.alignment = Alignment(vertical='center')

            # Color coding for status
            if col_idx == 9:  # Status column
                if material.status == 'Normal':
                    cell.fill = PatternFill('solid', fgColor='D1FAE5')
                    cell.font = Font(bold=True, color='065F46')
                elif material.status == 'Low Stock':
                    cell.fill = PatternFill('solid', fgColor='FEF9C3')
                    cell.font = Font(bold=True, color='78350F')
                elif material.status == 'Out of Stock':
                    cell.fill = PatternFill('solid', fgColor='FEE2E2')
                    cell.font = Font(bold=True, color='991B1B')

    # Summary sheet
    ws_summary = wb.create_sheet('Summary by Category')
    ws_summary.merge_cells('A1:D1')
    summary_title = ws_summary['A1']
    summary_title.value = f'Materials Summary by Category - {timezone.now().strftime("%B %Y")}'
    summary_title.font = Font(bold=True, size=14, color='1E1B4B')
    summary_title.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 30

    # Summary headers
    summary_headers = ['Category', 'Total Materials', 'In Stock', 'Low Stock', 'Out of Stock']
    ws_summary.append([])
    ws_summary.append(summary_headers)

    # Style summary header row
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=ws_summary.max_row, column=col_idx)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align
        cell.border = thin
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 18

    # Summary data by category
    from django.db.models import Count, Case, When, IntegerField
    category_summary = materials.values('category').annotate(
        total_materials=Count('id'),
        in_stock=Count(Case(When(status='Normal', then=1), output_field=IntegerField())),
        low_stock=Count(Case(When(status='Low Stock', then=1), output_field=IntegerField())),
        out_of_stock=Count(Case(When(status='Out of Stock', then=1), output_field=IntegerField())),
    ).order_by('category')

    for summary in category_summary:
        ws_summary.append([
            summary['category'],
            summary['total_materials'],
            summary['in_stock'],
            summary['low_stock'],
            summary['out_of_stock'],
        ])

        # Style summary data row
        for col_idx in range(1, 6):
            cell = ws_summary.cell(row=ws_summary.max_row, column=col_idx)
            cell.border = thin
            cell.alignment = Alignment(vertical='center')

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'materials_inventory_{timezone.now().strftime("%Y_%m_%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)
    return response


@login_required
def material_json(request, pk):
    """Return material data as JSON for populating the
     edit form via AJAX."""
    try:
        mat = Material.objects.get(pk=pk)
    except Material.DoesNotExist:
        return JsonResponse({'error': 'Material not found'}, status=404)

    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'
    # Only Storekeeper can fetch material data for edit form
    if role != 'Storekeeper':
        return JsonResponse({'error': 'Permission denied'}, status=403)

    data = {
        'id': mat.id,
        'name': mat.name,
        'category': mat.category,
        'quantity': mat.quantity,
        'rate': mat.rate,
        'Remaining_stock': mat.Remaining_stock,
        'min_stock_level': mat.min_stock_level,
        'Type': mat.Type,
    }
    return JsonResponse(data)

# @login_required
# def tasks_view(request):
#     profile = ensure_userprofile(request.user)
#     role = profile.role if profile else 'Branch'

#     if role == 'Branch':
#         tasks = Task.objects.filter(technician=request.user).order_by('-created_at')
#     else:
#         tasks = Task.objects.all().order_by('-created_at')

#     if request.method == 'POST':
#         action = request.POST.get('action')
        
#         if action == 'create':
#             if role == 'Branch':
#                  messages.error(request, "Branch users cannot create tasks.")
#                  return redirect('tasks')
#             form = TaskForm(request.POST)
#             if form.is_valid():
#                 form.save()
#                 messages.success(request, "Task created!")
#                 return redirect('tasks')
        
#         elif action == 'update_status':
#             task_id = request.POST.get('task_id')
#             new_status = request.POST.get('status')
#             try:
#                 task = Task.objects.get(pk=task_id)
#                 # Permission check
#                 if role == 'Branch' and task.requester != request.user:
#                     messages.error(request, "Permission denied.")
#                 else:
#                     task.status = new_status
#                     task.save()
#                     messages.success(request, f"Task status updated to {new_status}")
#             except Task.DoesNotExist:
#                 messages.error(request, "Task not found.")
#             return redirect('tasks')

#         elif action == 'delete':
#             if role != 'Admin':
#                 messages.error(request, "Only Admins can delete tasks.")
#                 return redirect('tasks')
#             task_id = request.POST.get('task_id')
#             try:
#                 task = Task.objects.get(pk=task_id)
#                 task.delete()
#                 messages.success(request, "Task deleted.")
#             except Task.DoesNotExist:
#                 messages.error(request, "Task not found.")
#             return redirect('tasks')

#     else:
#         form = TaskForm()
        
#     return render(request, 'inventory/tasks.html', {'tasks': tasks.order_by('-created_at'), 'form': form, 'role': role})

@login_required
def requests_view(request):
    now = timezone.now()
    base_requests = MaterialRequest.objects.filter(
        requester=request.user,
        requested_at__year=now.year,
        requested_at__month=now.month,
        is_archived=False
    ).order_by('-requested_at')
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'
    
    if role == 'NOC':
        return redirect('noc:dashboard')
    
    # For Admin/Storekeeper, show relevant requests instead of just their own
    if role in ['Admin', 'Storekeeper']:
        base_requests = MaterialRequest.objects.filter(
            requested_at__year=now.year,
            requested_at__month=now.month,
            is_archived=False,
            is_hidden_by_admin=False
        ).order_by('-requested_at')
    

    #Request count (pending/approved/rejected)
    pending_count = base_requests.filter(status='Pending').count()
    approved_count = base_requests.filter(status='Approved').count()
    dispatched_count = base_requests.filter(status='Dispatched').count()
    received_count = base_requests.filter(status='Received').count()
    rejected_count = base_requests.filter(status='Rejected').count()
    advance_count = base_requests.filter(request_type='Advance').count()
    
    # Get users for dropdown - only Branch role users
    users = User.objects.filter(userprofile__role='Branch').select_related('userprofile').annotate(
        request_count=Sum(
            Case(
                When(material_requests__status='Received', then=1),
                default=0,
                output_field=IntegerField()
            )
        )
    ).order_by('first_name', 'last_name')
    
    # Ensure all users have UserProfile
    for user in users:
        try:
            ensure_userprofile(user)
        except Exception:
            pass
    
    # Handle user dropdown filter - filter requests by selected branch user
    selected_user = None
    selected_user_id = request.GET.get('user', '').strip()
    if selected_user_id and selected_user_id.isdigit():
        try:
            selected_user = User.objects.get(id=selected_user_id)
            # Filter requests by this user
            base_requests = base_requests.filter(requester=selected_user)
            pending_count = base_requests.filter(status='Pending').count()
            approved_count = base_requests.filter(status='Approved').count()
            dispatched_count = base_requests.filter(status='Dispatched').count()
            received_count = base_requests.filter(status='Received').count()
            rejected_count = base_requests.filter(status='Rejected').count()
        except User.DoesNotExist:
            selected_user = None

    # Search Logic - apply BEFORE pagination
    search_query = request.GET.get('search', '').strip()
    if search_query:
        base_requests = base_requests.filter(
            Q(material__name__icontains=search_query) | 
            Q(send_by__icontains=search_query) | 
            Q(notes__icontains=search_query)|
            #type of request search
            Q(request_type__icontains=search_query)
            # Q(requester__username__icontains=search_query)
        )

    # Combine all requests (both Regular and Advance) for unified table display
    all_requests = base_requests.order_by('-requested_at')
    
    # Pagination applied to all requests combined
    paginator = Paginator(all_requests, 20)  # Show 20 requests per page
    page_number = request.GET.get('page')
    requests_page = paginator.get_page(page_number)
    
    # Count advance requests for display
    advance_count = base_requests.filter(request_type='Advance').count()

    # Initialize form - will be used in both GET and POST
    form = RequestForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax') == '1'

        # Save received_by (Branch requester only) via AJAX (expects JSON)
        if action == 'save_received_by':
            if role != 'Branch':
                return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

            req_id = (request.POST.get('req_id') or '').strip()
            received_by = (request.POST.get('received_by') or '').strip()

            if not req_id.isdigit():
                return JsonResponse({'success': False, 'error': 'Invalid request id.'}, status=400)
            if not received_by:
                return JsonResponse({'success': False, 'error': 'Received By is required.'}, status=400)

            try:
                req = MaterialRequest.objects.get(pk=int(req_id), requester=request.user)
            except MaterialRequest.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Request not found.'}, status=404)

            if req.status == 'Dispatched':
                pass
            elif req.status == 'Approved' and hasattr(req.material.created_by, 'userprofile') and req.material.created_by.userprofile.role == 'NOC':
                pass
            else:
                return JsonResponse({'success': False, 'error': 'Only dispatched requests or NOC-approved requests can be received.'}, status=400)

            req.received_by = received_by
            req.received_at = timezone.now()
            req.status = 'Received'
            req.save(update_fields=['received_by', 'received_at', 'status'])
            return JsonResponse({
                'success': True,
                'received_by': req.received_by,
                'received_at': req.received_at.isoformat() if req.received_at else None,
                'status': req.status,
            })

        # Save pass_on (Storekeeper/Admin only) via AJAX (expects JSON)
        if action == 'save_pass_on':
            if role not in ['Admin', 'Storekeeper']:
                return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

            req_id = (request.POST.get('req_id') or '').strip()
            pass_on = (request.POST.get('pass_on') or '').strip()

            if not req_id.isdigit():
                return JsonResponse({'success': False, 'error': 'Invalid request id.'}, status=400)
            if not pass_on:
                return JsonResponse({'success': False, 'error': 'Pass On details are required.'}, status=400)

            try:
                req = MaterialRequest.objects.get(pk=int(req_id))
            except MaterialRequest.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Request not found.'}, status=404)

            if req.status != 'Approved':
                return JsonResponse({'success': False, 'error': 'Only approved requests can be dispatched.'}, status=400)

            req.pass_on = pass_on
            req.pass_on_at = timezone.now()
            req.status = 'Dispatched'
            req.save(update_fields=['pass_on', 'pass_on_at', 'status'])
            return JsonResponse({
                'success': True,
                'pass_on': req.pass_on,
                'pass_on_at': req.pass_on_at.isoformat() if req.pass_on_at else None,
                'status': req.status,
            })
        
        # Create Request
        if action == 'create':
            if role in ['Admin', 'Storekeeper']:
                 messages.error(request, "Only Branch users can submit requests.")
                 return redirect('requests')

            form = RequestForm(request.POST)
            if form.is_valid():
                req = form.save(commit=False)
                req.requester = request.user
                
                # Get request_type from POST (comes from hidden field in form)
                request_type = request.POST.get('request_type', 'Regular')
                if request_type in ['Regular', 'Advance']:
                    req.request_type = request_type
                else:
                    req.request_type = 'Regular'
                
                req.save()
                req_type_display = 'Advance' if req.request_type == 'Advance' else 'Regular'
                messages.success(request, f"{req_type_display} request submitted successfully!")
                return redirect('requests')
        
        # Manage Request (Admin only)
        elif action in ['accept', 'reject', 'save_note', 'delete']:
            if role not in ['Admin', 'Storekeeper']:
                messages.error(request, "Permission denied.")
                return redirect('requests')
                
            req_id = request.POST.get('req_id')
            
            # Delete Action
            if action == 'delete':
                try:
                    req = MaterialRequest.objects.get(pk=req_id)
                    if req.status == 'Received':
                        req.is_hidden_by_admin = True
                        req.save()
                        messages.success(request, "Request hidden from your view successfully (Branch data preserved).")
                    elif req.status == 'Approved':
                        # Return stock before deleting
                        try:
                            with transaction.atomic():
                                mat = Material.objects.select_for_update().get(pk=req.material.id)
                                mat.quantity += req.quantity
                                mat.save()
                                req.delete()
                                messages.success(request, f"Approved request deleted. {req.quantity} units returned to {mat.name}.")
                        except Exception as e:
                            messages.error(request, f"Internal error during stock return: {str(e)}")
                    else:
                        req.delete()
                        messages.success(request, "Request deleted successfully.")
                except MaterialRequest.DoesNotExist:
                    messages.error(request, "Request not found.")
                return redirect('requests')

            note = request.POST.get('admin_note', '')
            admin_quantity = request.POST.get('quantity', '').strip()
            
            try:
                req = MaterialRequest.objects.get(pk=req_id)
                
                if action == 'accept':
                    if req.status in ['Approved', 'Dispatched', 'Received']:
                        messages.warning(request, "Request already approved.")
                        return redirect('requests')
                    
                    # Get the quantity to approve (admin can override)
                    try:
                        if admin_quantity:
                            approved_qty = int(admin_quantity)
                            if approved_qty <= 0:
                                messages.error(request, "Quantity must be greater than 0.")
                                return redirect('requests')
                        else:
                            # Use requested quantity if admin didn't specify
                            approved_qty = req.quantity
                    except ValueError:
                        messages.error(request, "Invalid quantity value.")
                        return redirect('requests')
                    
                    try:
                        with transaction.atomic():
                            # Refresh material to be safe
                            mat = Material.objects.select_for_update().get(pk=req.material.id)
                            total_available = mat.quantity + mat.Remaining_stock
                            
                            # Check if sufficient stock available
                            if approved_qty > total_available:
                                messages.error(request, f"Insufficient stock for {mat.name}. Available In Stock: {mat.quantity}, Remaining Stock: {mat.Remaining_stock}, Requested: {approved_qty}")
                                return redirect('requests')
                            
                            # Deduct the approved quantity (Prioritize In Stock / Quantity)
                            take_from_qty = min(approved_qty, mat.quantity)
                            take_from_rem = approved_qty - take_from_qty
                            
                            mat.quantity -= take_from_qty
                            mat.Remaining_stock -= take_from_rem
                            mat.save()
                            
                            # Update request with tracking info
                            req.quantity = approved_qty
                            req.deducted_from_quantity = take_from_qty
                            req.deducted_from_remaining = take_from_rem
                            req.status = 'Approved'
                            req.admin_note = note
                            req.save()
                            messages.success(request, f"Request approved. {approved_qty} units deducted (In Stock: {take_from_qty}, Remaining: {take_from_rem}).")
                    except Exception as e:
                         messages.error(request, f"Transaction failed: {str(e)}")
                         return redirect('requests')

                elif action == 'reject':
                    if req.status in ['Approved', 'Dispatched', 'Received']:
                        try:
                            with transaction.atomic():
                                # Return quantity to exactly where it was taken from
                                mat = Material.objects.select_for_update().get(pk=req.material.id)
                                mat.quantity += req.deducted_from_quantity
                                mat.Remaining_stock += req.deducted_from_remaining
                                mat.save()
                                
                                # Update request status and reset tracking
                                req.status = 'Rejected'
                                req.admin_note = note
                                req.deducted_from_quantity = 0
                                req.deducted_from_remaining = 0
                                req.save()
                                messages.success(request, f"Request rejected. Values returned to original stock pools.")
                        except Exception as e:
                             messages.error(request, f"Failed to return stock: {str(e)}")
                             return redirect('requests')
                    else:
                        req.status = 'Rejected'
                        req.admin_note = note
                        req.save()
                        messages.success(request, "Request rejected.")
                
                elif action == 'save_note':
                     req.admin_note = note
                     req.save()
                     messages.success(request, "Note saved.")

            except MaterialRequest.DoesNotExist:
                messages.error(request, "Request not found.")
            return redirect('requests')

    else:
        form = RequestForm()


    return render(request, 'inventory/requests.html', {
        'pending_count': pending_count,
        'approved_count': approved_count,
        'dispatched_count': dispatched_count,
        'received_count': received_count,
        'rejected_count': rejected_count,
        'requests': requests_page, 
        'form': form,
        'role': role,
        'page_obj': requests_page,
        'users': users,
        'advance_count': advance_count,
        'selected_user': selected_user,
    })

@login_required
def reports_view(request):
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return redirect('noc:dashboard')

    # ── Date range ──
    preset    = request.GET.get('preset', '')
    from_date = request.GET.get('from_date', '')
    to_date   = request.GET.get('to_date', '')
    report_type = request.GET.get('type', 'all')

    now = timezone.now()
    if preset == 'today':
        from_date = to_date = now.strftime('%Y-%m-%d')
    elif preset == 'week':
        from_date = (now - timezone.timedelta(days=6)).strftime('%Y-%m-%d')
        to_date   = now.strftime('%Y-%m-%d')
    elif preset == 'month':
        from_date = now.replace(day=1).strftime('%Y-%m-%d')
        to_date   = now.strftime('%Y-%m-%d')
    elif preset == 'last30':
        from_date = (now - timezone.timedelta(days=30)).strftime('%Y-%m-%d')
        to_date   = now.strftime('%Y-%m-%d')
    elif preset == 'last90':
        from_date = (now - timezone.timedelta(days=90)).strftime('%Y-%m-%d')
        to_date   = now.strftime('%Y-%m-%d')
    else:
        # Default: last 30 days if nothing provided
        if not from_date:
            from_date = (now - timezone.timedelta(days=30)).strftime('%Y-%m-%d')
        if not to_date:
            to_date = now.strftime('%Y-%m-%d')

    try:
        start = datetime.strptime(from_date, '%Y-%m-%d').date()
        end   = datetime.strptime(to_date,   '%Y-%m-%d').date()
    except ValueError:
        start = (now - timezone.timedelta(days=30)).date()
        end   = now.date()
        from_date = start.strftime('%Y-%m-%d')
        to_date   = end.strftime('%Y-%m-%d')

    # ── Base queryset ─────────────────────────────────────────────────────────
    requests_qs = MaterialRequest.objects.filter(
        requested_at__date__gte=start,
        requested_at__date__lte=end
    ).select_related('material', 'requester')

    # Role filter: Branch users see only their own requests
    if role == 'Branch':
        requests_qs = requests_qs.filter(requester=request.user)

    # ── Summary Stats ─────────────────────────────────────────────────────────
    total_requests   = requests_qs.count()
    approved_count   = requests_qs.filter(status='Received').count()
    pending_count    = requests_qs.filter(status='Pending').count()
    rejected_count   = requests_qs.filter(status='Rejected').count()
    total_qty_issued = requests_qs.filter(status='Received').aggregate(total=Sum('quantity'))['total'] or 0
    advance_count    = requests_qs.filter(request_type='Advance', status='Received').count()

    # Material stock summary
    total_materials  = Material.objects.count()
    low_stock_items  = Material.objects.filter(status='Low Stock').count()
    out_of_stock     = Material.objects.filter(status='Out of Stock').count()
    normal_stock     = Material.objects.filter(status='Normal').count()

    # Used materials in period
    used_qs = UsedMaterial.objects.filter(
        added_at__date__gte=start,
        added_at__date__lte=end
    )
    if role == 'Branch':
        used_qs = used_qs.filter(technician=request.user)
    total_used_records = used_qs.count()
    total_used_qty     = used_qs.aggregate(total=Sum('quantity'))['total'] or 0

    # ── Top 10 materials by approved quantity ─────────────────────────────────
    top_materials = (
        requests_qs.filter(status='Received')
        .values('material__name')
        .annotate(total_qty=Sum('quantity'), req_count=Count('id'))
        .order_by('-total_qty')[:10]
    )

    # ── Per-user breakdown (Admin/Storekeeper only) ───────────────────────────
    user_breakdown = []
    if role in ['Admin', 'Storekeeper']:
        user_breakdown = (
            requests_qs
            .values('requester__username', 'requester__first_name', 'requester__last_name')
            .annotate(
                total_req=Count('id'),
                approved=Count('id', filter=Q(status='Received')),
                pending=Count('id', filter=Q(status='Pending')),
                rejected=Count('id', filter=Q(status='Rejected')),
                qty_issued=Sum('quantity', filter=Q(status='Received'))
            ).order_by('-approved')[:15]
        )

    # ── Chart data: daily request counts over the date range ─────────────────
    daily_data = (
        requests_qs
        .annotate(day=TruncDate('requested_at'))
        .values('day')
        .annotate(
            approved=Count('id', filter=Q(status='Received')),
            pending=Count('id', filter=Q(status='Pending')),
            rejected=Count('id', filter=Q(status='Rejected')),
        )
        .order_by('day')
    )
    chart_labels   = [str(d['day']) for d in daily_data]
    chart_approved = [d['approved'] for d in daily_data]
    chart_pending  = [d['pending']  for d in daily_data]
    chart_rejected = [d['rejected'] for d in daily_data]

    # Material category breakdown
    category_data = (
        requests_qs.filter(status='Received')
        .values('material__category')
        .annotate(qty=Sum('quantity'))
        .order_by('-qty')
    )
    cat_labels = [d['material__category'] or 'Unknown' for d in category_data]
    cat_values = [d['qty'] or 0 for d in category_data]

    # ── Recent requests (up to 20 for table) ─────────────────────────────────
    recent_requests = requests_qs.order_by('-requested_at')[:20]

    # ── Low-stock materials list ──────────────────────────────────────────────
    low_stock_list = Material.objects.filter(
        status__in=['Low Stock', 'Out of Stock']
    ).order_by('status', 'name')[:20]

    # ── Branch-specific detailed data ─────────────────────────────────────────
    branch_req_pending  = []
    branch_req_approved = []
    branch_req_rejected = []
    branch_um_pending   = []
    branch_um_accepted  = []
    branch_um_rejected  = []

    if role == 'Branch':
        branch_req_pending  = requests_qs.filter(status='Pending').order_by('-requested_at')
        branch_req_approved = requests_qs.filter(status='Received').order_by('-requested_at')
        branch_req_rejected = requests_qs.filter(status='Rejected').order_by('-requested_at')

        branch_um_pending  = used_qs.filter(status='Pending').order_by('-added_at')
        branch_um_accepted = used_qs.filter(status='Accepted').order_by('-added_at')
        branch_um_rejected = used_qs.filter(status='Rejected').order_by('-added_at')

    # Used materials stats by status
    used_pending_count  = used_qs.filter(status='Pending').count()
    used_accepted_count = used_qs.filter(status='Accepted').count()
    used_rejected_count = used_qs.filter(status='Rejected').count()

    context = {
        # Date range
        'from_date':   from_date,
        'to_date':     to_date,
        'preset':      preset,
        'report_type': report_type,
        'role':        role,
        # Summary
        'total_requests':   total_requests,
        'approved_count':   approved_count,
        'pending_count':    pending_count,
        'rejected_count':   rejected_count,
        'total_qty_issued': total_qty_issued,
        'advance_count':    advance_count,
        'total_used_records': total_used_records,
        'total_used_qty':   total_used_qty,
        'used_pending_count':  used_pending_count,
        'used_accepted_count': used_accepted_count,
        'used_rejected_count': used_rejected_count,
        # Stock summary
        'total_materials': total_materials,
        'low_stock_items': low_stock_items,
        'out_of_stock':    out_of_stock,
        'normal_stock':    normal_stock,
        # Tables
        'top_materials':    top_materials,
        'user_breakdown':   user_breakdown,
        'recent_requests':  recent_requests,
        'low_stock_list':   low_stock_list,
        # Branch detailed tables
        'branch_req_pending':  branch_req_pending,
        'branch_req_approved': branch_req_approved,
        'branch_req_rejected': branch_req_rejected,
        'branch_um_pending':   branch_um_pending,
        'branch_um_accepted':  branch_um_accepted,
        'branch_um_rejected':  branch_um_rejected,
        # Chart data (serialised for JS)
        'chart_labels_json':   _json.dumps(chart_labels),
        'chart_approved_json': _json.dumps(chart_approved),
        'chart_pending_json':  _json.dumps(chart_pending),
        'chart_rejected_json': _json.dumps(chart_rejected),
        'cat_labels_json':     _json.dumps(cat_labels),
        'cat_values_json':     _json.dumps(cat_values),
    }
    return render(request, 'inventory/reports.html', context)


def _generate_branch_excel_report(request, requests_qs, start, end, from_date, to_date):
    """Generate Excel report for Branch users with separate sheets for each status."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    # Get used materials data
    used_qs = UsedMaterial.objects.filter(
        technician=request.user,
        added_at__date__gte=start,
        added_at__date__lte=end
    ).select_related('material', 'technician').order_by('-added_at')
    
    wb = openpyxl.Workbook()
    # Remove the default active sheet if it exists
    if wb.active:
        wb.remove(wb.active)
    
    # Style definitions
    h_fill   = PatternFill('solid', fgColor='4F46E5')
    h_font   = Font(color='FFFFFF', bold=True, size=11)
    h_align  = Alignment(horizontal='center', vertical='center')
    thin     = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    green  = PatternFill('solid', fgColor='D1FAE5')
    yellow = PatternFill('solid', fgColor='FEF9C3')
    red    = PatternFill('solid', fgColor='FEE2E2')
    
    def style_header_row(ws, headers, col_widths):
        ws.append(headers)
        for col_idx, width in enumerate(col_widths, 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill    = h_fill
            cell.font    = h_font
            cell.alignment = h_align
            cell.border  = thin
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def style_data_rows(ws, start_row):
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), 1):
            fill = PatternFill('solid', fgColor='F5F3FF') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            for cell in row:
                cell.border    = thin
                cell.fill      = fill
                cell.alignment = Alignment(vertical='center')
    
    # ── MATERIAL REQUESTS SHEETS ──────────────────────────────────────────────
    
    # Sheet 1: Pending Requests
    ws_req_pending = wb.create_sheet('Requests - Pending')
    ws_req_pending.row_dimensions[1].height = 22
    ws_req_pending.merge_cells('A1:G1')
    title = ws_req_pending['A1']
    title.value = f'Material Requests - PENDING ({from_date} → {to_date})'
    title.font = Font(bold=True, size=13, color='1E1B4B')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_req_pending.row_dimensions[1].height = 28
    ws_req_pending.append([])
    
    pending_requests = requests_qs.filter(status='Pending').order_by('-requested_at')
    headers = ['Date', 'Material', 'Category', 'Qty', 'Type', 'Notes', 'Status']
    widths  = [14, 28, 16, 8, 12, 30, 12]
    style_header_row(ws_req_pending, headers, widths)
    
    for req in pending_requests:
        ws_req_pending.append([
            req.requested_at.strftime('%Y-%m-%d'),
            req.material.name,
            req.material.category,
            req.quantity,
            req.request_type,
            req.notes or '',
            'Pending',
        ])
    style_data_rows(ws_req_pending, start_row=4)
    
    for row in ws_req_pending.iter_rows(min_row=4, max_row=ws_req_pending.max_row):
        row[6].fill = yellow
        row[6].font = Font(bold=True, color='78350F')
    
    # Sheet 2: Approved Requests
    ws_req_approved = wb.create_sheet('Requests - Approved')
    ws_req_approved.row_dimensions[1].height = 22
    ws_req_approved.merge_cells('A1:G1')
    title = ws_req_approved['A1']
    title.value = f'Material Requests - APPROVED ({from_date} → {to_date})'
    title.font = Font(bold=True, size=13, color='1E1B4B')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_req_approved.row_dimensions[1].height = 28
    ws_req_approved.append([])
    
    approved_requests = requests_qs.filter(status='Received').order_by('-requested_at')
    style_header_row(ws_req_approved, headers, widths)
    
    for req in approved_requests:
        ws_req_approved.append([
            req.requested_at.strftime('%Y-%m-%d'),
            req.material.name,
            req.material.category,
            req.quantity,
            req.request_type,
            req.notes or '',
            'Approved',
        ])
    style_data_rows(ws_req_approved, start_row=4)
    
    for row in ws_req_approved.iter_rows(min_row=4, max_row=ws_req_approved.max_row):
        row[6].fill = green
        row[6].font = Font(bold=True, color='065F46')
    
    # Sheet 3: Rejected Requests
    ws_req_rejected = wb.create_sheet('Requests - Rejected')
    ws_req_rejected.row_dimensions[1].height = 22
    ws_req_rejected.merge_cells('A1:G1')
    title = ws_req_rejected['A1']
    title.value = f'Material Requests - REJECTED ({from_date} → {to_date})'
    title.font = Font(bold=True, size=13, color='1E1B4B')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_req_rejected.row_dimensions[1].height = 28
    ws_req_rejected.append([])
    
    rejected_requests = requests_qs.filter(status='Rejected').order_by('-requested_at')
    style_header_row(ws_req_rejected, headers, widths)
    
    for req in rejected_requests:
        ws_req_rejected.append([
            req.requested_at.strftime('%Y-%m-%d'),
            req.material.name,
            req.material.category,
            req.quantity,
            req.request_type,
            req.notes or '',
            'Rejected',
        ])
    style_data_rows(ws_req_rejected, start_row=4)
    
    for row in ws_req_rejected.iter_rows(min_row=4, max_row=ws_req_rejected.max_row):
        row[6].fill = red
        row[6].font = Font(bold=True, color='991B1B')
    
    # ── USED MATERIALS SHEETS ─────────────────────────────────────────────────
    
    # Sheet 4: Pending Used Materials
    ws_um_pending = wb.create_sheet('Used Materials - Pending')
    ws_um_pending.row_dimensions[1].height = 22
    ws_um_pending.merge_cells('A1:F1')
    title = ws_um_pending['A1']
    title.value = f'Used Materials - PENDING ({from_date} → {to_date})'
    title.font = Font(bold=True, size=13, color='1E1B4B')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_um_pending.row_dimensions[1].height = 28
    ws_um_pending.append([])
    
    um_pending = used_qs.filter(status='Pending').order_by('-added_at')
    um_headers = ['Date', 'Material', 'Category', 'Qty Used', 'Notes', 'Status']
    um_widths  = [14, 28, 16, 12, 30, 12]
    style_header_row(ws_um_pending, um_headers, um_widths)
    
    for um in um_pending:
        ws_um_pending.append([
            um.added_at.strftime('%Y-%m-%d'),
            um.material.name,
            um.material.category,
            um.quantity,
            um.issue or '',
            'Pending',
        ])
    style_data_rows(ws_um_pending, start_row=4)
    
    for row in ws_um_pending.iter_rows(min_row=4, max_row=ws_um_pending.max_row):
        row[5].fill = yellow
        row[5].font = Font(bold=True, color='78350F')
    
    # Sheet 5: Accepted Used Materials
    ws_um_accepted = wb.create_sheet('Used Materials - Accepted')
    ws_um_accepted.row_dimensions[1].height = 22
    ws_um_accepted.merge_cells('A1:F1')
    title = ws_um_accepted['A1']
    title.value = f'Used Materials - ACCEPTED ({from_date} → {to_date})'
    title.font = Font(bold=True, size=13, color='1E1B4B')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_um_accepted.row_dimensions[1].height = 28
    ws_um_accepted.append([])
    
    um_accepted = used_qs.filter(status='Accepted').order_by('-added_at')
    style_header_row(ws_um_accepted, um_headers, um_widths)
    
    for um in um_accepted:
        ws_um_accepted.append([
            um.added_at.strftime('%Y-%m-%d'),
            um.material.name,
            um.material.category,
            um.quantity,
            um.issue or '',
            'Accepted',
        ])
    style_data_rows(ws_um_accepted, start_row=4)
    
    for row in ws_um_accepted.iter_rows(min_row=4, max_row=ws_um_accepted.max_row):
        row[5].fill = green
        row[5].font = Font(bold=True, color='065F46')
    
    # Sheet 6: Rejected Used Materials
    ws_um_rejected = wb.create_sheet('Used Materials - Rejected')
    ws_um_rejected.row_dimensions[1].height = 22
    ws_um_rejected.merge_cells('A1:F1')
    title = ws_um_rejected['A1']
    title.value = f'Used Materials - REJECTED ({from_date} → {to_date})'
    title.font = Font(bold=True, size=13, color='1E1B4B')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_um_rejected.row_dimensions[1].height = 28
    ws_um_rejected.append([])
    
    um_rejected = used_qs.filter(status='Rejected').order_by('-added_at')
    style_header_row(ws_um_rejected, um_headers, um_widths)
    
    for um in um_rejected:
        ws_um_rejected.append([
            um.added_at.strftime('%Y-%m-%d'),
            um.material.name,
            um.material.category,
            um.quantity,
            um.issue or '',
            'Rejected',
        ])
    style_data_rows(ws_um_rejected, start_row=4)
    
    for row in ws_um_rejected.iter_rows(min_row=4, max_row=ws_um_rejected.max_row):
        row[5].fill = red
        row[5].font = Font(bold=True, color='991B1B')
    
    # ── Summary Sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet('Summary', 0)  # Insert at beginning
    ws_summary.row_dimensions[1].height = 28
    ws_summary.merge_cells('A1:D1')
    title = ws_summary['A1']
    title.value = f'Report Summary ({from_date} → {to_date})'
    title.font = Font(bold=True, size=14, color='1E1B4B')
    title.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_summary.append([])
    
    # Summary statistics
    req_pending_count = pending_requests.count()
    req_approved_count = approved_requests.count()
    req_rejected_count = rejected_requests.count()
    um_pending_count = um_pending.count()
    um_accepted_count = um_accepted.count()
    um_rejected_count = um_rejected.count()
    
    req_approved_qty = approved_requests.aggregate(total=Sum('quantity'))['total'] or 0
    um_accepted_qty = um_accepted.aggregate(total=Sum('quantity'))['total'] or 0
    
    summary_data = [
        [],
        ['Material Requests', ''],
        ['Status', 'Count', 'Qty'],
        ['Pending', req_pending_count, '-'],
        ['Approved', req_approved_count, req_approved_qty],
        ['Rejected', req_rejected_count, '-'],
        [],
        ['Used Materials', ''],
        ['Status', 'Count', 'Qty'],
        ['Pending', um_pending_count, '-'],
        ['Accepted', um_accepted_count, um_accepted_qty],
        ['Rejected', um_rejected_count, '-'],
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    # Style summary sheet
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    
    for row in ws_summary.iter_rows(min_row=3, max_row=5):
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='E0E7FF')
            cell.font = Font(bold=True)
            cell.border = thin
    
    for row in ws_summary.iter_rows(min_row=8, max_row=11):
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='E0E7FF')
            cell.font = Font(bold=True)
            cell.border = thin
    
    # Freeze panes
    for ws in [ws_summary, ws_req_pending, ws_req_approved, ws_req_rejected, ws_um_pending, ws_um_accepted, ws_um_rejected]:
        ws.freeze_panes = 'A4'
    
    # Save and return
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"branch_report_{from_date}_to_{to_date}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def reports_export_excel(request):
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return redirect('noc:dashboard')

    from_date = request.GET.get('from_date', (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d'))
    to_date   = request.GET.get('to_date',   timezone.now().strftime('%Y-%m-%d'))
    try:
        start = datetime.strptime(from_date, '%Y-%m-%d').date()
        end   = datetime.strptime(to_date,   '%Y-%m-%d').date()
    except ValueError:
        start = (timezone.now() - timezone.timedelta(days=30)).date()
        end   = timezone.now().date()

    requests_qs = MaterialRequest.objects.filter(
        requested_at__date__gte=start,
        requested_at__date__lte=end
    ).select_related('material', 'requester').order_by('-requested_at')
    if role == 'Branch':
        requests_qs = requests_qs.filter(requester=request.user)

    # ── Branch Role: Specialized Report with Separated Status Sheets ──
    if role == 'Branch':
        return _generate_branch_excel_report(request, requests_qs, start, end, from_date, to_date)

    wb = openpyxl.Workbook()

    # ── Helper styles ─────────────────────────────────────────────────────────
    h_fill   = PatternFill('solid', fgColor='4F46E5')
    h_font   = Font(color='FFFFFF', bold=True, size=11)
    h_align  = Alignment(horizontal='center', vertical='center')
    thin     = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    alt_fill = PatternFill('solid', fgColor='F5F3FF')

    def style_header_row(ws, headers, col_widths):
        ws.append(headers)
        for col_idx, width in enumerate(col_widths, 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill    = h_fill
            cell.font    = h_font
            cell.alignment = h_align
            cell.border  = thin
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def style_data_rows(ws, start_row):
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), 1):
            fill = PatternFill('solid', fgColor='F5F3FF') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            for cell in row:
                cell.border    = thin
                cell.fill      = fill
                cell.alignment = Alignment(vertical='center')

    # ── Sheet 1: Material Requests ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Material Requests'
    ws1.row_dimensions[1].height = 22

    # Title row
    ws1.merge_cells('A1:G1')
    title_cell = ws1['A1']
    title_cell.value     = f'ISP Inventory — Material Requests Report  ({from_date}  →  {to_date})'
    title_cell.font      = Font(bold=True, size=13, color='1E1B4B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28
    ws1.append([])  # blank

    headers = ['Date', 'Requester', 'Material', 'Category', 'Qty', 'Type', 'Status']
    widths  = [14, 22, 28, 16, 8, 12, 12]
    style_header_row(ws1, headers, widths)

    for req in requests_qs:
        ws1.append([
            req.requested_at.strftime('%Y-%m-%d'),
            req.requester.get_full_name() or req.requester.username,
            req.material.name,
            req.material.category,
            req.quantity,
            req.request_type,
            req.status,
        ])
    style_data_rows(ws1, start_row=4)

    # Status colour coding
    green  = PatternFill('solid', fgColor='D1FAE5')
    yellow = PatternFill('solid', fgColor='FEF9C3')
    red    = PatternFill('solid', fgColor='FEE2E2')
    for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row):
        status_cell = row[6]
        if status_cell.value == 'Approved':
            status_cell.fill = green
            status_cell.font = Font(bold=True, color='065F46')
        elif status_cell.value == 'Pending':
            status_cell.fill = yellow
            status_cell.font = Font(bold=True, color='78350F')
        elif status_cell.value == 'Rejected':
            status_cell.fill = red
            status_cell.font = Font(bold=True, color='991B1B')

    # ── Sheet 2: Top Materials ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Top Materials')
    ws2.merge_cells('A1:C1')
    ws2['A1'].value = 'Top Materials by Approved Qty'
    ws2['A1'].font  = Font(bold=True, size=12, color='1E1B4B')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 24
    ws2.append([])
    style_header_row(ws2, ['Material', 'Category', 'Total Approved Qty'], [30, 18, 22])
    top = (
        requests_qs.filter(status='Approved')
        .values('material__name', 'material__category')
        .annotate(total_qty=Sum('quantity'))
        .order_by('-total_qty')[:20]
    )
    for item in top:
        ws2.append([item['material__name'], item['material__category'], item['total_qty']])
    style_data_rows(ws2, start_row=4)

    # ── Sheet 3: User Summary ─────────────────────────────────────────────────
    ws3 = wb.create_sheet('User Summary')
    ws3.merge_cells('A1:F1')
    ws3['A1'].value = 'Per-User Request Summary'
    ws3['A1'].font  = Font(bold=True, size=12, color='1E1B4B')
    ws3['A1'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 24
    ws3.append([])
    style_header_row(ws3, ['Username', 'Full Name', 'Total Requests', 'Approved', 'Pending', 'Qty Issued'], [18, 24, 16, 12, 12, 14])
    user_data = (
        requests_qs
        .values('requester__username', 'requester__first_name', 'requester__last_name')
        .annotate(
            total_req=Count('id'),
            approved=Count('id', filter=Q(status='Approved')),
            pending=Count('id', filter=Q(status='Pending')),
            qty_issued=Sum('quantity', filter=Q(status='Approved'))
        )
        .order_by('-approved')
    )
    for u in user_data:
        fn = f"{u['requester__first_name']} {u['requester__last_name']}".strip() or u['requester__username']
        ws3.append([
            u['requester__username'], fn,
            u['total_req'], u['approved'], u['pending'],
            u['qty_issued'] or 0,
        ])
    style_data_rows(ws3, start_row=4)

    # ── Sheet 4: Stock Status ─────────────────────────────────────────────────
    ws4 = wb.create_sheet('Stock Status')
    ws4.merge_cells('A1:D1')
    ws4['A1'].value = 'Current Stock Status'
    ws4['A1'].font  = Font(bold=True, size=12, color='1E1B4B')
    ws4['A1'].alignment = Alignment(horizontal='center')
    ws4.row_dimensions[1].height = 24
    ws4.append([])
    style_header_row(ws4, ['Material', 'Category', 'Quantity', 'Status'], [30, 18, 12, 14])
    for mat in Material.objects.order_by('status', 'name'):
        ws4.append([mat.name, mat.category, mat.quantity, mat.status])
    style_data_rows(ws4, start_row=4)
    for row in ws4.iter_rows(min_row=4, max_row=ws4.max_row):
        sc = row[3]
        if sc.value == 'Normal':
            sc.fill = green;  sc.font = Font(bold=True, color='065F46')
        elif sc.value == 'Low Stock':
            sc.fill = yellow; sc.font = Font(bold=True, color='78350F')
        elif sc.value == 'Out of Stock':
            sc.fill = red;    sc.font = Font(bold=True, color='991B1B')

    # ── Freeze top rows & return ──────────────────────────────────────────────
    for ws in [ws1, ws2, ws3, ws4]:
        ws.freeze_panes = 'A4'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"isp_report_{from_date}_to_{to_date}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _generate_branch_pdf_report(request, requests_qs, start, end, from_date, to_date):
    """Generate PDF report for Branch users with requests and used materials."""
    from xhtml2pdf import pisa
    from io import BytesIO
    
    # Get used materials data
    used_qs = UsedMaterial.objects.filter(
        technician=request.user,
        added_at__date__gte=start,
        added_at__date__lte=end
    ).select_related('material', 'technician').order_by('-added_at')
    
    # Calculate statistics
    req_pending = requests_qs.filter(status='Pending')
    req_approved = requests_qs.filter(status='Approved')
    req_rejected = requests_qs.filter(status='Rejected')
    
    um_pending = used_qs.filter(status='Pending')
    um_accepted = used_qs.filter(status='Accepted')
    um_rejected = used_qs.filter(status='Rejected')
    
    context = {
        'from_date': from_date,
        'to_date': to_date,
        'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
        'generated_by': request.user.get_full_name() or request.user.username,
        'branch_name': request.user.get_full_name() or request.user.username,
        # Request Statistics
        'req_pending_count': req_pending.count(),
        'req_approved_count': req_approved.count(),
        'req_rejected_count': req_rejected.count(),
        'req_total_count': requests_qs.count(),
        'req_approved_qty': req_approved.aggregate(total=Sum('quantity'))['total'] or 0,
        # Used Materials Statistics
        'um_pending_count': um_pending.count(),
        'um_accepted_count': um_accepted.count(),
        'um_rejected_count': um_rejected.count(),
        'um_total_count': used_qs.count(),
        'um_accepted_qty': um_accepted.aggregate(total=Sum('quantity'))['total'] or 0,
        # Detailed data
        'req_pending_list': req_pending[:50],
        'req_approved_list': req_approved[:50],
        'req_rejected_list': req_rejected[:50],
        'um_pending_list': um_pending[:50],
        'um_accepted_list': um_accepted[:50],
        'um_rejected_list': um_rejected[:50],
    }
    
    html_string = render(request, 'inventory/branch_report_pdf.html', context).content.decode('utf-8')
    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('PDF generation error', status=500)
    buffer.seek(0)
    filename = f"branch_report_{from_date}_to_{to_date}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def reports_export_pdf(request):
    """Export the current report as a PDF file using xhtml2pdf."""
    from xhtml2pdf import pisa
    from io import BytesIO
    from django.db.models import Count

    profile = ensure_userprofile(request.user)
    role    = profile.role if profile else 'Branch'

    if role == 'NOC':
        return redirect('noc:dashboard')

    from_date = request.GET.get('from_date', (timezone.now() - timezone.timedelta(days=30)).strftime('%Y-%m-%d'))
    to_date   = request.GET.get('to_date',   timezone.now().strftime('%Y-%m-%d'))
    try:
        start = datetime.strptime(from_date, '%Y-%m-%d').date()
        end   = datetime.strptime(to_date,   '%Y-%m-%d').date()
    except ValueError:
        start = (timezone.now() - timezone.timedelta(days=30)).date()
        end   = timezone.now().date()

    requests_qs = MaterialRequest.objects.filter(
        requested_at__date__gte=start,
        requested_at__date__lte=end
    ).select_related('material', 'requester').order_by('-requested_at')
    if role == 'Branch':
        requests_qs = requests_qs.filter(requester=request.user)
        return _generate_branch_pdf_report(request, requests_qs, start, end, from_date, to_date)

    # ── Admin/Storekeeper Report ──
    total_requests   = requests_qs.count()
    approved_count   = requests_qs.filter(status='Approved').count()
    pending_count    = requests_qs.filter(status='Pending').count()
    rejected_count   = requests_qs.filter(status='Rejected').count()
    total_qty_issued = requests_qs.filter(status='Approved').aggregate(total=Sum('quantity'))['total'] or 0

    top_materials = (
        requests_qs.filter(status='Approved')
        .values('material__name', 'material__category')
        .annotate(total_qty=Sum('quantity'))
        .order_by('-total_qty')[:15]
    )

    context = {
        'from_date': from_date, 'to_date': to_date,
        'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
        'generated_by': request.user.get_full_name() or request.user.username,
        'total_requests': total_requests, 'approved_count': approved_count,
        'pending_count': pending_count, 'rejected_count': rejected_count,
        'total_qty_issued': total_qty_issued,
        'requests_qs': requests_qs[:100],
        'top_materials': top_materials,
        'low_stock_list': Material.objects.filter(status__in=['Low Stock', 'Out of Stock']).order_by('status', 'name'),
    }

    html_string = render(request, 'inventory/report_pdf.html', context).content.decode('utf-8')
    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('PDF generation error', status=500)
    buffer.seek(0)
    filename = f"isp_report_{from_date}_to_{to_date}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def settings_view(request):
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return redirect('noc:dashboard')

    # Ensure role groups exist
    ROLE_GROUPS = ['Admin', 'Storekeeper', 'Branch', 'NOC']
    for r in ROLE_GROUPS:
        Group.objects.get_or_create(name=r)

    # Admin access: allow either UserProfile role==Admin or membership in Admin group
    try:
        is_admin_profile = (request.user.userprofile.role == 'Admin')
    except Exception:
        is_admin_profile = False

    is_admin_group = request.user.groups.filter(name='Admin').exists()
    if not (is_admin_profile or is_admin_group):
        messages.error(request, "Only Admins can access Settings!")
        return redirect('dashboard')

    # Use User queryset for compatibility with existing template which expects User objects
    users = User.objects.all().select_related('userprofile')
    default_group = Group.objects.get(name='Branch')
    # Ensure every user has a UserProfile and at least one role-group
    for u in users:
        # create UserProfile if missing, prefilling role from first role-group if available
        try:
            ensure_userprofile(u)
        except Exception:
            # best effort; continue
            pass
        # ensure at least one role-group assigned
        if not u.groups.filter(name__in=ROLE_GROUPS).exists():
            u.groups.add(default_group)
    system_settings = SystemSetting.objects.all()

    # Notification form for current user
    notif_obj, _ = NotificationSetting.objects.get_or_create(user=request.user)
    notif_form = NotificationSettingForm(instance=notif_obj)

    setting_form = SystemSettingForm()

    if request.method == 'POST':
        action = request.POST.get('action')

        # ── Create User ──
        if action == 'create_user':
            username = request.POST.get('username', '').strip()
            email = request.POST.get('email', '').strip()
            password = request.POST.get('password', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            role = request.POST.get('role', 'Branch')
            phone = request.POST.get('phone', '').strip()
            address = request.POST.get('address', '').strip()
            city = request.POST.get('city', '').strip()
            zip_code = request.POST.get('zip_code', '').strip()

            if not username or not password:
                messages.error(request, "Username and password are required.")
                return redirect('settings')

            # Validate password strength
            try:
                validate_password(password)
            except ValidationError as e:
                for msg in e.messages:
                    messages.error(request, msg)
                return redirect('settings')

            if User.objects.filter(username=username).exists():
                messages.error(request, f"Username '{username}' already exists.")
                return redirect('settings')

            try:
                new_user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                )
                # Assign role group
                grp, _ = Group.objects.get_or_create(name=role)
                new_user.groups.add(grp)
                # Create UserProfile
                profile, _ = UserProfile.objects.get_or_create(user=new_user)
                profile.role = role
                profile.phone = phone
                profile.address = address
                profile.city = city
                profile.zip_code = zip_code
                image = request.FILES.get('image')
                if image:
                    profile.image = image
                profile.save()
                messages.success(request, f"User '{username}' created successfully!")
            except Exception as e:
                messages.error(request, f"Error creating user: {str(e)}")
            return redirect('settings')

        # ── Edit User ──────────────────────────────────────────────────────
        elif action == 'edit_user':
            user_id = request.POST.get('user_id')
            if not user_id:
                messages.error(request, "Invalid user.")
                return redirect('settings')
            try:
                edit_user = User.objects.get(id=user_id)
                # Update User fields
                new_username = request.POST.get('username', '').strip()
                new_email = request.POST.get('email', '').strip()
                new_first_name = request.POST.get('first_name', '').strip()
                new_last_name = request.POST.get('last_name', '').strip()
                new_role = request.POST.get('role', '')
                new_phone = request.POST.get('phone', '').strip()
                new_address = request.POST.get('address', '').strip()
                new_city = request.POST.get('city', '').strip()
                new_zip_code = request.POST.get('zip_code', '').strip()
                is_active = request.POST.get('is_active') == 'on'
                new_password = request.POST.get('password', '').strip()

                # Check username uniqueness (exclude current user)
                if new_username and new_username != edit_user.username:
                    if User.objects.filter(username=new_username).exclude(id=user_id).exists():
                        messages.error(request, f"Username '{new_username}' is already taken.")
                        return redirect('settings')
                    edit_user.username = new_username

                if new_email:
                    edit_user.email = new_email
                edit_user.first_name = new_first_name
                edit_user.last_name = new_last_name
                edit_user.is_active = is_active
                if new_password:
                    # Validate new password strength
                    try:
                        validate_password(new_password, edit_user)
                        edit_user.set_password(new_password)
                    except ValidationError as e:
                        for msg in e.messages:
                            messages.error(request, msg)
                        return redirect('settings')
                edit_user.save()

                # Update role group
                if new_role and new_role in ROLE_GROUPS:
                    for rn in ROLE_GROUPS:
                        g = Group.objects.filter(name=rn).first()
                        if g and g in edit_user.groups.all():
                            edit_user.groups.remove(g)
                    grp, _ = Group.objects.get_or_create(name=new_role)
                    edit_user.groups.add(grp)

                # Update UserProfile
                profile, _ = UserProfile.objects.get_or_create(user=edit_user)
                if new_role:
                    profile.role = new_role
                profile.phone = new_phone
                profile.address = new_address
                profile.city = new_city
                profile.zip_code = new_zip_code
                new_image = request.FILES.get('image')
                if new_image:
                    profile.image = new_image
                profile.save()

                messages.success(request, f"User '{edit_user.username}' updated successfully!")
            except User.DoesNotExist:
                messages.error(request, "User not found.")
            except Exception as e:
                messages.error(request, f"Error updating user: {str(e)}")
            return redirect('settings')

        # ── Toggle Active Status ───────────────────────────────────────────
        elif action == 'toggle_status':
            user_id = request.POST.get('user_id')
            if user_id:
                try:
                    toggle_user = User.objects.get(id=user_id)
                    if toggle_user.is_superuser:
                        messages.error(request, "Cannot deactivate superuser accounts.")
                    else:
                        toggle_user.is_active = not toggle_user.is_active
                        toggle_user.save()
                        status_text = "activated" if toggle_user.is_active else "deactivated"
                        messages.success(request, f"User '{toggle_user.username}' {status_text}.")
                except User.DoesNotExist:
                    messages.error(request, "User not found.")
            return redirect('settings')

        # ── Delete User (supports both 'delete' and 'delete_user') ────────
        elif action in ['delete', 'delete_user']:
            user_id = request.POST.get('user_id')
            if user_id:
                try:
                    del_user = User.objects.get(id=user_id)
                    if del_user.is_superuser:
                        messages.error(request, "Cannot delete superuser accounts.")
                    elif del_user == request.user:
                        messages.error(request, "You cannot delete your own account.")
                    else:
                        del_username = del_user.username
                        del_user.delete()
                        messages.success(request, f"User '{del_username}' deleted successfully.")
                except User.DoesNotExist:
                    messages.error(request, "User not found.")
            return redirect('settings')

        elif action == 'update_notifications':
            form = NotificationSettingForm(request.POST, instance=notif_obj)
            if form.is_valid():
                form.save()
                messages.success(request, "Notification preferences updated successfully!")
            else:
                messages.error(request, "Error updating notification preferences.")
            return redirect('settings')
        
        elif action == 'update_logs':
            # Admin only
            user_role = profile.role
            if user_role != 'Admin':
                messages.error(request, "Only Admin can update log settings.")
                return redirect('settings')
            
            try:
                log_settings, created = LogSettings.objects.get_or_create(pk=1)
                log_settings.log_level = request.POST.get('log_level', 'INFO')
                log_settings.enable_file_logging = request.POST.get('enable_file_logging') == 'on'
                log_settings.enable_database_logging = request.POST.get('enable_database_logging') == 'on'
                log_settings.log_user_activities = request.POST.get('log_user_activities') == 'on'
                log_settings.updated_by = request.user
                log_settings.save()
                messages.success(request, "Log settings updated successfully!")
            except Exception as e:
                messages.error(request, f"Error updating log settings: {str(e)}")
            return redirect('settings')

        elif action == 'backup':
            """Create backup with full or partial data support"""
            import json
            import hashlib
            import tempfile
            import os
            from django.core.files.base import ContentFile
            from io import StringIO
            
            # Check if user is Admin or can create backups (Storekeeper, Branch, NOC)
            user_role = profile.role
            if user_role not in ['Admin', 'Storekeeper', 'Branch', 'NOC']:
                messages.error(request, "You don't have permission to create backups.")
                return redirect('settings')
            
            try:
                backup_type = request.POST.get('backup_type', 'full')
                description = request.POST.get('backup_description', '').strip()
                
                # Determine models to backup based on backup type
                if backup_type == 'full':
                    # Full backup: all data including users
                    exclude_models = ['auth.permission', 'contenttypes', 'admin.logentry']
                    backup_label = 'Full Backup'
                else:
                    # Partial backup: only materials, requests, and usage data (no user accounts)
                    exclude_models = [
                        'auth.permission', 'auth.user', 'auth.group', 
                        'contenttypes', 'admin.logentry', 'sessions.session'
                    ]
                    backup_label = 'Partial Backup (Data Only)'
                
                # Create backup data using dumpdata
                output = StringIO()
                call_command('dumpdata', exclude=exclude_models, stdout=output, indent=2)
                backup_data = output.getvalue()
                
                if not backup_data:
                    messages.error(request, "Backup failed: No data to backup.")
                    return redirect('settings')
                
                # Calculate checksum for integrity verification
                checksum = hashlib.sha256(backup_data.encode()).hexdigest()
                
                # Count records
                try:
                    data_dict = json.loads(backup_data)
                    records_count = len(data_dict) if isinstance(data_dict, list) else len(data_dict)
                except json.JSONDecodeError:
                    records_count = 0
                
                # Save backup to model
                from isp_inventory.models import BackupRestore
                
                timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f'backup_{backup_type}_{timestamp}.json'
                
                backup_file = ContentFile(
                    backup_data.encode(),
                    name=backup_filename
                )
                
                backup_obj = BackupRestore(
                    backup_file=backup_file,
                    created_by=request.user,
                    backup_type=backup_type,
                    backup_size=len(backup_data),
                    description=description or f'{backup_label} created by {request.user.username}',
                    status='active',
                    data_records_count=records_count,
                    checksum=checksum,
                )
                backup_obj.save()
                
                # Return download response
                response = HttpResponse(backup_data, content_type='application/json')
                response['Content-Disposition'] = f'attachment; filename="{backup_filename}"'
                messages.success(request, f"{backup_label} created successfully! Records: {records_count}, Size: {backup_obj.get_file_size_display()}")
                return response
                
            except Exception as e:
                import traceback
                messages.error(request, f"Backup failed: {str(e)}")
                print(f"Backup Error: {traceback.format_exc()}")
                return redirect('settings')
        
        elif action == 'restore':
            """Restore from backup with full or partial data support"""
            # Only Admin can restore
            user_role = profile.role
            if user_role != 'Admin':
                messages.error(request, "Only Admin can restore backups.")
                return redirect('settings')
            
            try:
                import json
                import tempfile
                import os
                from isp_inventory.models import BackupRestore
                
                # Check if restoring from file upload or existing backup
                restore_type = request.POST.get('restore_type', 'file')
                backup_content = None
                backup_obj = None
                
                if restore_type == 'file' and 'backup_file' in request.FILES:
                    # Restore from uploaded file
                    uploaded_file = request.FILES['backup_file']
                    backup_content = uploaded_file.read().decode('utf-8')
                    
                elif restore_type == 'history':
                    # Restore from backup history
                    backup_id = request.POST.get('backup_id')
                    if not backup_id:
                        messages.error(request, "Please select a backup to restore.")
                        return redirect('settings')
                    
                    backup_obj = BackupRestore.objects.filter(id=backup_id, status__in=['active', 'deleted']).first()
                    if not backup_obj:
                        messages.error(request, "Backup not found or not recoverable.")
                        return redirect('settings')
                    
                    backup_content = backup_obj.backup_file.read().decode('utf-8')
                else:
                    messages.error(request, "Invalid restore type.")
                    return redirect('settings')
                
                # Parse and validate JSON
                backup_data = json.loads(backup_content)
                if not backup_data:
                    messages.error(request, "Backup file is empty.")
                    return redirect('settings')
                
                # Confirmation required
                confirm = request.POST.get('confirm_restore')
                if confirm != 'yes':
                    messages.warning(request, "Please confirm restoration to proceed.")
                    return redirect('settings')
                
                # Determine restore type
                is_partial = backup_obj and backup_obj.backup_type == 'partial'
                
                # Perform restore with proper transaction handling
                with transaction.atomic():
                    # Create temporary file for loaddata
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
                        json.dump(backup_data, tmp, indent=2)
                        tmp_path = tmp.name
                    
                    try:
                        if is_partial:
                            # Partial restore: clear only data tables, keep users and auth
                            from django.db import connection
                            from django.apps import apps
                            
                            cursor = connection.cursor()
                            
                            # Tables to clear for partial restore (data only)
                            tables_to_clear = [
                                'isp_inventory_material',
                                'isp_inventory_materialrequest',
                                'isp_inventory_usedmaterial',
                                'isp_inventory_macserialnum',
                                'isp_inventory_refundablematerial',
                                'isp_inventory_damagematerial',
                                'isp_inventory_materialmonthlycoun',
                                'isp_inventory_backuprestore',
                                'isp_inventory_activitylog',
                            ]
                            
                            for table in tables_to_clear:
                                try:
                                    cursor.execute(f'DELETE FROM {table}')
                                except Exception:
                                    pass  # Table might not exist
                            
                            connection.commit()
                        else:
                            # Full restore: clear all data
                            call_command('flush', '--no-input')
                        
                        # Load backup data
                        call_command('loaddata', tmp_path)
                        
                        # Update backup object if restoring from history
                        if backup_obj and restore_type == 'history':
                            backup_obj.restored_at = timezone.now()
                            backup_obj.restored_by = request.user
                            backup_obj.save()
                        
                        restore_type_label = 'Partial' if is_partial else 'Full'
                        messages.success(request, f"{restore_type_label} restore completed successfully! {len(backup_data)} records restored.")
                        
                    finally:
                        # Clean up temporary file
                        try:
                            os.unlink(tmp_path)
                        except Exception:
                            pass
                
            except json.JSONDecodeError:
                messages.error(request, "Invalid backup file format (not valid JSON).")
                return redirect('settings')
            except Exception as e:
                import traceback
                messages.error(request, f"Restore failed: {str(e)}")
                print(f"Restore Error: {traceback.format_exc()}")
                return redirect('settings')
        
        elif action == 'delete_backup':
            """Soft delete a backup (recoverable for 30 days)"""
            from isp_inventory.models import BackupRestore
            
            backup_id = request.POST.get('backup_id')
            try:
                backup = BackupRestore.objects.get(id=backup_id)
                user_role = profile.role
                
                # Authorization check
                if user_role == 'Admin':
                    # Admin can delete any backup
                    pass
                elif user_role in ['Storekeeper', 'Branch', 'NOC']:
                    # Can only delete own backups
                    if backup.created_by != request.user:
                        messages.error(request, "You can only delete your own backups.")
                        return redirect('settings')
                else:
                    messages.error(request, "You don't have permission to delete backups.")
                    return redirect('settings')
                
                # Soft delete
                backup.status = 'deleted'
                backup.deleted_at = timezone.now()
                backup.deleted_by = request.user
                backup.save()
                
                messages.success(request, "Backup moved to trash (recoverable for 30 days).")
                
            except BackupRestore.DoesNotExist:
                messages.error(request, "Backup not found.")
            except Exception as e:
                messages.error(request, f"Error deleting backup: {str(e)}")
            
            return redirect('settings')
        
        elif action == 'recover_backup':
            """Recover a soft-deleted backup"""
            from isp_inventory.models import BackupRestore
            
            backup_id = request.POST.get('backup_id')
            try:
                backup = BackupRestore.objects.get(id=backup_id, status='deleted')
                backup.status = 'active'
                backup.deleted_at = None
                backup.deleted_by = None
                backup.save()
                messages.success(request, "Backup recovered successfully.")
            except BackupRestore.DoesNotExist:
                messages.error(request, "Backup not found or already active.")
            except Exception as e:
                messages.error(request, f"Error recovering backup: {str(e)}")
            
            return redirect('settings')

        # Role change: update groups and (optionally) UserProfile for compatibility
        elif action == 'change_role':
            user_id = request.POST.get('user_id')
            new_role = request.POST.get('role') or request.POST.get('new_role')
            if user_id and new_role:
                user = User.objects.get(id=user_id)
                # ensure group exists
                grp, _ = Group.objects.get_or_create(name=new_role)
                # remove existing role groups
                for rn in ROLE_GROUPS:
                    g = Group.objects.filter(name=rn).first()
                    if g and g in user.groups.all():
                        user.groups.remove(g)
                user.groups.add(grp)
                
                # Update UserProfile role
                prof = ensure_userprofile(user)
                prof.role = new_role
                prof.save()
                messages.success(request, f"Role for '{user.username}' changed to {new_role}.")
            return redirect('settings')
        # Group management: create/delete groups, add/remove members
        elif action == 'create_group':
            group_name = request.POST.get('group_name', '').strip()
            if group_name:
                grp, created = Group.objects.get_or_create(name=group_name)
                if created:
                    messages.success(request, f"Group '{group_name}' created!")
                else:
                    messages.warning(request, f"Group '{group_name}' already exists.")
            else:
                messages.error(request, "Group name cannot be empty.")

        elif action == 'delete_group':
            group_id = request.POST.get('group_id')
            if group_id:
                try:
                    grp = Group.objects.get(id=group_id)
                    if grp.name in ROLE_GROUPS:
                        messages.error(request, "Cannot delete built-in role groups.")
                    else:
                        grp.delete()
                        messages.success(request, f"Group '{grp.name}' deleted!")
                except Group.DoesNotExist:
                    messages.error(request, "Group not found.")

        elif action == 'add_user_to_group':
            user_id = request.POST.get('user_id')
            group_id = request.POST.get('group_id')
            try:
                user = User.objects.get(id=user_id)
                grp = Group.objects.get(id=group_id)
                user.groups.add(grp)
                messages.success(request, f"User {user.username} added to group {grp.name}!")
            except (User.DoesNotExist, Group.DoesNotExist):
                messages.error(request, "User or group not found.")

        elif action == 'remove_user_from_group':
            user_id = request.POST.get('user_id')
            group_id = request.POST.get('group_id')
            try:
                user = User.objects.get(id=user_id)
                grp = Group.objects.get(id=group_id)
                user.groups.remove(grp)
                messages.success(request, f"User {user.username} removed from group {grp.name}!")
            except (User.DoesNotExist, Group.DoesNotExist):
                messages.error(request, "User or group not found.")


        return redirect('settings')

    # Notification tab values
    email_notifications = notif_obj.email_notifications
    in_app_notifications = notif_obj.in_app_notifications
    request_approved_alert = notif_obj.request_approved_alert
    request_rejected_alert = notif_obj.request_rejected_alert
    new_request_alert = notif_obj.new_request_alert
    low_stock_alert = notif_obj.low_stock_alert
    out_of_stock_alert = notif_obj.out_of_stock_alert
    material_destroyed_alert = notif_obj.material_destroyed_alert
    task_assignment_alert = notif_obj.task_assignment_alert
    task_completed_alert = notif_obj.task_completed_alert
    message_alert = notif_obj.message_alert
    backup_alert = notif_obj.backup_alert
    system_alert = notif_obj.system_alert

    # Log settings
    log_settings = LogSettings.objects.first()
    if not log_settings:
        log_settings = LogSettings.objects.create()
    log_level = log_settings.log_level
    enable_file_logging = log_settings.enable_file_logging
    enable_database_logging = log_settings.enable_database_logging
    log_user_activities = log_settings.log_user_activities
    
    # Get recent activity logs for current user
    recent_activity = ActivityLog.objects.filter(user=request.user).order_by('-timestamp')[:20]
    
    # Get backup history
    from isp_inventory.models import BackupRestore
    backup_history = BackupRestore.objects.all().select_related('created_by', 'deleted_by', 'restored_by').order_by('-created_at')[:20]
    
    context = {
        'users': users,
        'groups': Group.objects.all(),
        'system_settings': system_settings,
        'setting_form': setting_form,
        'notif_form': notif_form,
        'log_settings_form': LogSettingsForm(instance=log_settings),
        # Notification fields
        'email_notifications': email_notifications,
        'in_app_notifications': in_app_notifications,
        'request_approved_alert': request_approved_alert,
        'request_rejected_alert': request_rejected_alert,
        'new_request_alert': new_request_alert,
        'low_stock_alert': low_stock_alert,
        'out_of_stock_alert': out_of_stock_alert,
        'material_destroyed_alert': material_destroyed_alert,
        'task_assignment_alert': task_assignment_alert,
        'task_completed_alert': task_completed_alert,
        'message_alert': message_alert,
        'backup_alert': backup_alert,
        'system_alert': system_alert,
        # Log settings
        'log_level': log_level,
        'enable_file_logging': enable_file_logging,
        'enable_database_logging': enable_database_logging,
        'log_user_activities': log_user_activities,
        'log_settings': log_settings,
        'recent_activity': recent_activity,
        'backup_history': backup_history,
    }
    return render(request, 'inventory/settings.html', context)


@login_required
def profile_view(request):
    """View and update current user's profile."""
    profile = ensure_userprofile(request.user)
    role = profile.role

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        city = request.POST.get('city', '').strip()
        zip_code = request.POST.get('zip_code', '').strip()
        new_password = request.POST.get('password', '').strip()
        image = request.FILES.get('image')

        # Username editing restriction
        if role == 'Admin':
            username = request.POST.get('username', '').strip()
            if username and username != request.user.username:
                if User.objects.filter(username=username).exclude(id=request.user.id).exists():
                    messages.error(request, f"Username '{username}' is already taken.")
                else:
                    request.user.username = username
        
        request.user.email = email
        request.user.first_name = first_name
        request.user.last_name = last_name

        if new_password:
            try:
                validate_password(new_password, request.user)
                request.user.set_password(new_password)
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, request.user)
            except ValidationError as e:
                for msg in e.messages:
                    messages.error(request, msg)
                return redirect('profile')

        request.user.save()

        profile.phone = phone
        profile.address = address
        profile.city = city
        profile.zip_code = zip_code
        if image:
            # Simple validation for the manual profile update
            if image.size > 2 * 1024 * 1024:
                messages.error(request, "Profile image must be under 2 MB.")
                return redirect('profile')
            
            allowed = ('.png', '.jpg', '.jpeg', '.webp', '.gif')
            if not image.name.lower().endswith(allowed):
                messages.error(request, "Allowed image formats: PNG, JPG, JPEG, WEBP, GIF.")
                return redirect('profile')
                
            profile.image = image
        profile.save()

        messages.success(request, "Profile updated successfully!")
        return redirect('profile')

    return render(request, 'inventory/profile.html', {
        'profile': profile,
        'user': request.user,
    })


@login_required
def used_materials_view(request):
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return redirect('noc:dashboard')

    # Determine which used materials to display based on role
    if role == 'Branch':
        used_materials_qs = UsedMaterial.objects.filter(technician=request.user, is_archived=False).order_by('-added_at')
        branch_users = None
    else:
        # Admin/Storekeeper see all used materials
        used_materials_qs = UsedMaterial.objects.filter(is_archived=False).order_by('-added_at')
        branch_users = User.objects.select_related('userprofile').filter(userprofile__role='Branch').order_by('username')
        
        # Handle user dropdown filter
        selected_user_id = request.GET.get('user_id')
        if selected_user_id:
            try:
                selected_user = User.objects.select_related('userprofile').get(id=selected_user_id, userprofile__role='Branch')
                used_materials_qs = used_materials_qs.filter(technician=selected_user)
            except User.DoesNotExist:
                messages.error(request, "Selected user not found.")

    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        used_materials_qs = used_materials_qs.filter(
            Q(material__name__icontains=search_query) |
            Q(technician__username__icontains=search_query) |
            Q(technician__first_name__icontains=search_query) |
            Q(technician__last_name__icontains=search_query)|
            Q(client_name__icontains=search_query) |
            Q(client_phone__icontains=search_query) |
            Q(mac_serial__mac_serial__icontains=search_query)
        ).distinct()

    # Pagination - AFTER all filters are applied
    paginator = Paginator(used_materials_qs, 20)  # Show 20 records per page
    page_number = request.GET.get('page')
    used_materials_page = paginator.get_page(page_number)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        # BRANCH USER ACTIONS: create, edit, delete
        if action == 'create':
            if role != 'Branch':
                messages.error(request, "Only Branch users can add Used Materials.")
                return redirect('used_materials')
            
            form = UsedMaterialForm(request.POST, user=request.user)
            if form.is_valid():
                selection = form.cleaned_data.get('material_selection')
                prefix, pk = selection.split(':')
                
                if prefix == 's':
                    try:
                        mac = MacSerialNumber.objects.get(id=pk, assigned_to=request.user)
                        material = mac.material
                        mac_serial = mac
                        # Force quantity to 1 for serialized items
                        quantity = 1
                    except MacSerialNumber.DoesNotExist:
                        messages.error(request, "Selected Mac/Serial not found or not assigned to you.")
                        return redirect('used_materials')
                else:
                    try:
                        material = Material.objects.get(id=pk)
                        mac_serial = None
                        quantity = form.cleaned_data.get('quantity')
                    except Material.DoesNotExist:
                        messages.error(request, "Selected material not found.")
                        return redirect('used_materials')

                # Check if material is in approved stock
                approved_material_ids = MaterialRequest.objects.filter(
                    requester=request.user, status='Received'
                ).values_list('material', flat=True)
                
                if material.id in approved_material_ids:
                    um = form.save(commit=False)
                    um.technician = request.user
                    um.material = material
                    um.mac_serial = mac_serial
                    um.quantity = quantity
                    um.save()
                    
                    # Update MacSerial status based on status
                    if um.mac_serial:
                        if um.status == 'Accepted':
                            um.mac_serial.status = 'Used'
                        else:
                            um.mac_serial.status = 'Active'
                        um.mac_serial.save()
                        
                    messages.success(request, "Used Material recorded successfully!")
                    return redirect('used_materials')
                else:
                    messages.error(request, "You can only record usage for received materials.")
                    return redirect('used_materials')
            else:
                for field, errors in form.errors.items():
                    if errors:
                        messages.error(request, f"{field}: {errors[0]}")
                        break
                return redirect('used_materials')
                
        elif action == 'edit':
            if role != 'Branch':
                messages.error(request, "Only Branch users can edit used materials.")
                return redirect('used_materials')
                
            um_id = request.POST.get('um_id')
            try:
                um = UsedMaterial.objects.get(pk=um_id, technician=request.user)
                form = UsedMaterialForm(request.POST, instance=um, user=request.user)
                if form.is_valid():
                    selection = form.cleaned_data.get('material_selection')
                    prefix, pk = selection.split(':')
                    
                    if prefix == 's':
                        try:
                            mac = MacSerialNumber.objects.get(id=pk, assigned_to=request.user)
                            material = mac.material
                            new_mac = mac
                            new_qty = 1
                        except MacSerialNumber.DoesNotExist:
                            messages.error(request, "Selected Mac/Serial not found.")
                            return redirect('used_materials')
                    else:
                        try:
                            material = Material.objects.get(id=pk)
                            new_mac = None
                            new_qty = form.cleaned_data.get('quantity')
                        except Material.DoesNotExist:
                            messages.error(request, "Selected material not found.")
                            return redirect('used_materials')

                    approved_material_ids = MaterialRequest.objects.filter(
                        requester=request.user, status='Received'
                    ).values_list('material', flat=True).distinct()
                    
                    if material.id in approved_material_ids:
                        old_um = UsedMaterial.objects.get(pk=um.pk)
                        old_mac = old_um.mac_serial
                        new_status = form.cleaned_data.get('status')

                        try:
                            with transaction.atomic():
                                # Handle MacSerial return/deduction
                                if old_mac:
                                    old_mac.status = 'Active'
                                    old_mac.save()
                                
                                if new_mac:
                                    if new_status == 'Accepted':
                                        new_mac.status = 'Used'
                                    else:
                                        new_mac.status = 'Active'
                                    new_mac.save()
                                
                                um.material = material
                                um.mac_serial = new_mac
                                um.quantity = new_qty
                                um.status = new_status
                                um.save()
                                messages.success(request, "Used Material updated successfully.")
                        except Exception as e:
                            messages.error(request, f"Update error: {str(e)}")
                            return redirect('used_materials')
                        
                        return redirect('used_materials')
                    else:
                        messages.error(request, "You can only use approved materials.")
                else:
                    messages.error(request, "Invalid form data.")
            except UsedMaterial.DoesNotExist:
                messages.error(request, "Record not found.")
            return redirect('used_materials')
        
        elif action == 'delete':
            if role != 'Branch':
                messages.error(request, "Only Branch users can delete used materials.")
                return redirect('used_materials')
                
            um_id = request.POST.get('um_id')
            try:
                um = UsedMaterial.objects.get(pk=um_id, technician=request.user)
                
                # Return Mac/Serial if applicable
                if um.mac_serial:
                    um.mac_serial.status = 'Active'
                    um.mac_serial.save()
                
                um.delete()
                messages.success(request, "Used Material deleted.")
            except UsedMaterial.DoesNotExist:
                messages.error(request, "Record not found.")
            return redirect('used_materials')

        # BRANCH USER ACTIONS: accept, reject (Now allowed for Branch as requested)
        elif action == 'accept':
            um_id = request.POST.get('um_id')
            admin_note = request.POST.get('admin_note', '')
            try:
                used_material = UsedMaterial.objects.get(pk=um_id, technician=request.user)
                if used_material.status == 'Accepted':
                    messages.warning(request, "Already accepted.")
                else:
                    try:
                        with transaction.atomic():
                            if used_material.mac_serial:
                                used_material.mac_serial.status = 'Used'
                                used_material.mac_serial.save()
                            
                            used_material.status = 'Accepted'
                            used_material.admin_note = admin_note
                            used_material.save()
                            messages.success(request, "Usage confirmed.")
                    except Exception as e:
                        messages.error(request, f"Error: {str(e)}")
            except UsedMaterial.DoesNotExist:
                messages.error(request, "Record not found.")
            return redirect('used_materials')
            
        elif action == 'reject':
            um_id = request.POST.get('um_id')
            admin_note = request.POST.get('admin_note', '')
            try:
                used_material = UsedMaterial.objects.get(pk=um_id, technician=request.user)
                if used_material.status == 'Rejected':
                    messages.warning(request, "Already rejected.")
                else:
                    try:
                        with transaction.atomic():
                            if used_material.mac_serial:
                                used_material.mac_serial.status = 'Active'
                                used_material.mac_serial.save()
                                
                            messages.success(request, "Rejected and serial returned.")
                            
                            used_material.status = 'Rejected'
                            used_material.admin_note = admin_note
                            used_material.save()
                    except Exception as e:
                        messages.error(request, f"Error: {str(e)}")
            except UsedMaterial.DoesNotExist:
                messages.error(request, "Record not found.")
            return redirect('used_materials')

    else:
        form = UsedMaterialForm(user=request.user)

    # Determine selected_used_material and for_approval flags for modal display (Branch only now)
    selected_used_material = None
    for_approval = False
    um_id = request.GET.get('um_id')
    if um_id:
        try:
            selected_used_material = UsedMaterial.objects.get(pk=um_id, technician=request.user)
            for_approval = True
        except UsedMaterial.DoesNotExist:
            pass

    # Build serials dict for JS filtering
    all_user_serials = MacSerialNumber.objects.filter(
        assigned_to=request.user,
        status='Active'
    )
    serials_by_material = {}
    for s in all_user_serials:
        if s.material_id not in serials_by_material:
            serials_by_material[s.material_id] = []
        serials_by_material[s.material_id].append({'id': s.id, 'serial': s.mac_serial})

    return render(request, 'inventory/used_materials.html', {
        'used_materials': used_materials_page,
        'form': form,
        'role': role,
        'page_obj': used_materials_page,
        'selected_used_material': selected_used_material,
        'for_approval': for_approval,
        'branch_users': branch_users,
        'search_query': search_query,
        'serials_by_material_json': _json.dumps(serials_by_material),
    })

@login_required
def get_used_material_api(request, pk):
    """API endpoint to get used material data for editing via AJAX"""
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return JsonResponse({'error': 'NOC role restricted from this API.'}, status=403)

    try:
        if role in ['Admin', 'Storekeeper']:
            used_material = UsedMaterial.objects.get(pk=pk)
        else:
            used_material = UsedMaterial.objects.get(pk=pk, technician=request.user)
    except UsedMaterial.DoesNotExist:
        return JsonResponse({'error': 'Record not found or access denied'}, status=404)

    selection = f"s:{used_material.mac_serial.id}" if used_material.mac_serial else f"m:{used_material.material.id}"
    
    data = {
        'id': used_material.id,
        'material_selection': selection,
        'client_name': used_material.client_name or '',
        'client_phone': used_material.client_phone or '',
        'client_address': used_material.client_address or '',
        'quantity': used_material.quantity,
        'issue': used_material.issue or '',
        'status': used_material.status,
    }
    return JsonResponse(data)

@login_required
def manage_used_material_api(request, pk):
    """API endpoint to manage (approve/reject) used materials for admins"""
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    # Only Branch users can manage their own used materials
    if role != 'Branch':
        return JsonResponse({'error': 'Permission denied. Only Branch users can manage their usage.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        used_material = UsedMaterial.objects.get(pk=pk, technician=request.user)
    except UsedMaterial.DoesNotExist:
        return JsonResponse({'error': 'Record not found or access denied'}, status=404)

    action = request.POST.get('action')
    admin_note = request.POST.get('admin_note', '').strip()

    try:
        with transaction.atomic():
            if action == 'accept':
                if used_material.status == 'Accepted':
                    used_material.admin_note = admin_note
                    used_material.save()
                    return JsonResponse({'success': True, 'message': 'Status is already Accepted. Note updated.'})

                used_material.status = 'Accepted'
                used_material.admin_note = admin_note
                used_material.save()
                
                # Update Mac/Serial status to 'Used'
                if used_material.mac_serial:
                    used_material.mac_serial.status = 'Used'
                    used_material.mac_serial.save()
                
                return JsonResponse({'success': True, 'message': 'Usage confirmed and serial locked.'})

            elif action == 'reject':
                if used_material.status == 'Rejected':
                    used_material.admin_note = admin_note
                    used_material.save()
                    return JsonResponse({'success': True, 'message': 'Status is already Rejected. Note updated.'})

                # Return Mac/Serial status to 'Active'
                if used_material.mac_serial:
                    used_material.mac_serial.status = 'Active'
                    used_material.mac_serial.save()
                
                used_material.status = 'Rejected'
                used_material.admin_note = admin_note
                used_material.save()

                return JsonResponse({'success': True, 'message': 'Rejected and serial released.'})
            else:
                return JsonResponse({'error': 'Invalid action'}, status=400)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def pending_requests_api(request):
    """
    API endpoint to fetch pending requests with count and ordering.
    Returns JSON data with:
    - List of pending material requests ordered by requested_at (newest first)
    - Total count of pending requests
    - Count of pending requests by user
    """
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role == 'NOC':
        return JsonResponse({'error': 'NOC role restricted from this API.'}, status=403)
    
    try:
        # Get pending requests ordered by most recent first
        pending_requests = MaterialRequest.objects.filter(
            status='Pending',
            is_archived=False
        ).select_related('requester', 'material').order_by('-requested_at')
        
        # For non-admin users, optionally filter to their own requests
        show_all = request.GET.get('show_all', 'true').lower() == 'true'
        if not show_all and role == 'Branch':
            pending_requests = pending_requests.filter(requester=request.user)
        
        # Pagination
        page_size = int(request.GET.get('page_size', 10))
        page_number = int(request.GET.get('page', 1))
        
        paginator = Paginator(pending_requests, page_size)
        page_obj = paginator.get_page(page_number)
        
        # Build request data
        requests_data = []
        for req in page_obj:
            requests_data.append({
                'id': req.id,
                'requester': req.requester.get_full_name() or req.requester.username,
                'requester_username': req.requester.username,
                'material_name': req.material.name,
                'material_id': req.material.id,
                'quantity': req.quantity,
                'requested_at': req.requested_at.isoformat(),
                'requested_at_display': req.requested_at.strftime('%Y-%m-%d %H:%M'),
                'status': req.status,
                'notes': req.notes or '',
            })
        
        # Return JSON response with metadata
        return JsonResponse({
            'success': True,
            'data': requests_data,
            'count': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page_number,
            'page_size': page_size,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def chat_view(request):
    profile = ensure_userprofile(request.user)
    
    from django.db.models import Max, Q
    from django.utils import timezone
    from datetime import datetime

    # Get all users with a profile, excluding self
    # This ensures we have a profile to access roles
    users_qs = User.objects.filter(
        userprofile__isnull=False
    ).exclude(id=request.user.id).select_related('userprofile').annotate(
        last_sent=Max('sent_messages__created_at', filter=Q(sent_messages__receiver=request.user)),
        last_received=Max('received_messages__created_at', filter=Q(received_messages__sender=request.user))
    )
    
    users = list(users_qs)
    # Define a very old date for users with no chat history
    epoch = timezone.make_aware(datetime(1970, 1, 1))
    
    for u in users:
        # Determine the latest activity time
        activity_dates = [d for d in [u.last_sent, u.last_received] if d is not None]
        u.latest_activity = max(activity_dates) if activity_dates else epoch
        
        # Attach the actual last message object for the snippet
        u.last_message = InternalMessage.objects.filter(
            Q(sender=request.user, receiver=u) | Q(sender=u, receiver=request.user)
        ).order_by('-created_at').first()

    # Sort users by latest activity descending (most recent first)
    users.sort(key=lambda x: x.latest_activity, reverse=True)

    base_template = 'noc/base.html' if profile.role == 'NOC' else 'inventory/base.html'

    return render(request, 'inventory/chat.html', {
        'role': profile.role,
        'users': users,
        'user': request.user,
        'base_template': base_template,
    })

@login_required
def chat_history_api(request, user_id):
    """Fetch chat history between current user and target user."""
    target_user = get_object_or_404(User, id=user_id)
    messages = InternalMessage.objects.filter(
        Q(sender=request.user, receiver=target_user) | 
        Q(sender=target_user, receiver=request.user)
    ).order_by('created_at')
    
    # Mark received messages from this user as read
    messages.filter(receiver=request.user, is_read=False).update(is_read=True)
    
    messages_data = []
    for m in messages:
        messages_data.append({
            'id': m.id,
            'sender_id': m.sender.id,
            'content': m.content,
            'created_at': m.created_at.isoformat(),
            'is_me': m.sender_id == request.user.id
        })
    
    return JsonResponse({
        'success': True,
        'messages': messages_data,
        'target_user': {
            'id': target_user.id,
            'username': target_user.username,
            'full_name': target_user.get_full_name() or target_user.username
        }
    })

@login_required
def refundable_materials_view(request):
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    # Get refundable materials list
    if role in ['Admin', 'Storekeeper']:
        refundable_qs = RefundableMaterial.objects.select_related('branch_user').order_by('-added_at')
        branch_users = User.objects.select_related('userprofile').filter(userprofile__role='Branch').order_by('username')
        
        # Handle user dropdown filter
        selected_user_id = request.GET.get('user_id')
        if selected_user_id:
            try:
                selected_user = User.objects.select_related('userprofile').get(id=selected_user_id, userprofile__role='Branch')
                refundable_qs = refundable_qs.filter(branch_user=selected_user)
            except User.DoesNotExist:
                messages.error(request, "Selected user not found.")
    elif role == 'NOC':
        refundable_qs = RefundableMaterial.objects.filter(branch_user__userprofile__role='Branch').select_related('branch_user').order_by('-added_at')
        branch_users = User.objects.select_related('userprofile').filter(userprofile__role='Branch').order_by('username')

        selected_user_id = request.GET.get('user_id')
        if selected_user_id:
            try:
                selected_user = User.objects.select_related('userprofile').get(id=selected_user_id, userprofile__role='Branch')
                refundable_qs = refundable_qs.filter(branch_user=selected_user)
            except User.DoesNotExist:
                messages.error(request, "Selected user not found.")
    else:
        # Branch user
        refundable_qs = RefundableMaterial.objects.filter(branch_user=request.user).order_by('-added_at')
        branch_users = None

    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        if role in ['Admin', 'Storekeeper', 'NOC']:
            refundable_qs = refundable_qs.filter(
                Q(material_name__icontains=search_query) |
                Q(branch_user__username__icontains=search_query) |
                Q(branch_user__first_name__icontains=search_query) |
                Q(branch_user__last_name__icontains=search_query)
            ).distinct()
        else:
            refundable_qs = refundable_qs.filter(
                Q(material_name__icontains=search_query)
            ).distinct()

    # Pagination
    paginator = Paginator(refundable_qs, 20)
    page_number = request.GET.get('page')
    refundable_page = paginator.get_page(page_number)

    # Forms and usage list
    form = RefundableMaterialForm(user=request.user)
    usage_form = RefundableMaterialUsageForm(user=request.user)

    refundable_usages_qs = RefundableMaterialUsage.objects.select_related('refundable_material', 'used_by').order_by('-used_at')
    if role == 'Branch':
        refundable_usages_qs = refundable_usages_qs.filter(used_by=request.user)
    elif role == 'NOC':
        refundable_usages_qs = refundable_usages_qs.filter(refundable_material__branch_user__userprofile__role='Branch')

    usage_paginator = Paginator(refundable_usages_qs, 20)
    usage_page = usage_paginator.get_page(request.GET.get('usage_page'))

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            if role != 'Branch':
                messages.error(request, "Only Branch users can add Refundable Materials.")
                return redirect('refundable_materials')

            form = RefundableMaterialForm(request.POST, user=request.user)
            if form.is_valid():
                rf = form.save(commit=False)
                rf.branch_user = request.user
                rf.save()
                messages.success(request, "Refundable Material logged successfully!")
                return redirect('refundable_materials')
            else:
                for field, errors in form.errors.items():
                    if errors:
                        messages.error(request, f"{field}: {errors[0]}")
                        break
                return redirect('refundable_materials')

        elif action == 'create_usage':
            if role != 'Branch':
                messages.error(request, "Only Branch users can add Used Materials.")
                return redirect('refundable_materials')

            usage_form = RefundableMaterialUsageForm(request.POST, user=request.user)
            if usage_form.is_valid():
                usage_form.save()
                messages.success(request, "Used material entry recorded successfully!")
                return redirect('refundable_materials')
            else:
                for field, errors in usage_form.errors.items():
                    if errors:
                        messages.error(request, f"{field}: {errors[0]}")
                        break
                return redirect('refundable_materials')

        elif action == 'edit':
            if role != 'Branch':
                messages.error(request, "Only Branch users can edit Refundable Materials.")
                return redirect('refundable_materials')

            rf_id = request.POST.get('rf_id')
            try:
                rf = RefundableMaterial.objects.get(pk=rf_id, branch_user=request.user)
                form = RefundableMaterialForm(request.POST, instance=rf, user=request.user)
                if form.is_valid():
                    form.save()
                    messages.success(request, "Refundable Material updated successfully!")
                    return redirect('refundable_materials')
                else:
                    for field, errors in form.errors.items():
                        if errors:
                            messages.error(request, f"{field}: {errors[0]}")
                            break
                    return redirect('refundable_materials')
            except RefundableMaterial.DoesNotExist:
                messages.error(request, "Record not found or access denied.")
                return redirect('refundable_materials')

        elif action == 'delete':
            if role != 'Branch':
                messages.error(request, "Only Branch users can delete Refundable Materials.")
                return redirect('refundable_materials')

            rf_id = request.POST.get('rf_id')
            try:
                rf = RefundableMaterial.objects.get(pk=rf_id, branch_user=request.user)
                rf.delete()
                messages.success(request, "Refundable Material record deleted.")
                return redirect('refundable_materials')
            except RefundableMaterial.DoesNotExist:
                messages.error(request, "Record not found or access denied.")
                return redirect('refundable_materials')

        elif action == 'delete_usage':
            if role != 'Branch':
                messages.error(request, "Only Branch users can delete Used Materials.")
                return redirect('refundable_materials')

            usage_id = request.POST.get('usage_id')
            try:
                usage = RefundableMaterialUsage.objects.get(pk=usage_id, used_by=request.user)
                usage.delete()
                messages.success(request, "Used material record deleted.")
            except RefundableMaterialUsage.DoesNotExist:
                messages.error(request, "Record not found or access denied.")
            return redirect('refundable_materials')

    template_name = 'noc/refundable_materials.html' if role == 'NOC' else 'inventory/refundable_materials.html'
    return render(request, template_name, {
        'refundable_materials': refundable_page,
        'form': form,
        'usage_form': usage_form,
        'usage_page': usage_page,
        'role': role,
        'page_obj': refundable_page,
        'branch_users': branch_users,
        'search_query': search_query,
    })

@login_required
def get_refundable_material_api(request, pk):
    """API endpoint to get refundable material data for editing via AJAX"""
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    try:
        if role in ['Admin', 'Storekeeper']:
            rf = RefundableMaterial.objects.get(pk=pk)
        elif role == 'NOC':
            rf = RefundableMaterial.objects.filter(branch_user__userprofile__role='Branch').get(pk=pk)
        else:
            rf = RefundableMaterial.objects.get(pk=pk, branch_user=request.user)
    except RefundableMaterial.DoesNotExist:
        return JsonResponse({'error': 'Record not found or access denied'}, status=404)

    data = {
        'id': rf.id,
        'material_name': rf.material_name,
        'quantity': rf.quantity,
    }
    return JsonResponse(data)

@login_required
def damaged_materials_view(request):
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    # Get damaged materials list
    if role in ['Admin', 'Storekeeper']:
        damaged_qs = DamageMaterial.objects.select_related('branch_user', 'material', 'confirmed_by').order_by('-added_at')
        branch_users = User.objects.select_related('userprofile').filter(userprofile__role='Branch').order_by('username')
        
        # Handle user dropdown filter
        selected_user_id = request.GET.get('user_id')
        if selected_user_id:
            try:
                selected_user = User.objects.select_related('userprofile').get(id=selected_user_id, userprofile__role='Branch')
                damaged_qs = damaged_qs.filter(branch_user=selected_user)
            except User.DoesNotExist:
                messages.error(request, "Selected user not found.")
    elif role == 'NOC':
        damaged_qs = DamageMaterial.objects.filter(material__category='Internet', material__created_by=request.user).select_related('branch_user', 'material', 'confirmed_by').order_by('-added_at')
        branch_users = User.objects.select_related('userprofile').filter(userprofile__role='Branch').order_by('username')
        
        # Handle user dropdown filter
        selected_user_id = request.GET.get('user_id')
        if selected_user_id:
            try:
                selected_user = User.objects.select_related('userprofile').get(id=selected_user_id, userprofile__role='Branch')
                damaged_qs = damaged_qs.filter(branch_user=selected_user)
            except User.DoesNotExist:
                messages.error(request, "Selected user not found.")
    else:
        # Branch user
        damaged_qs = DamageMaterial.objects.filter(branch_user=request.user).select_related('material').order_by('-added_at')
        branch_users = None

    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        if role in ['Admin', 'Storekeeper', 'NOC']:
            damaged_qs = damaged_qs.filter(
                Q(material__name__icontains=search_query) |
                Q(branch_user__username__icontains=search_query) |
                Q(branch_user__first_name__icontains=search_query) |
                Q(branch_user__last_name__icontains=search_query) |
                Q(damage_reason__icontains=search_query)
            ).distinct()
        else:
            damaged_qs = damaged_qs.filter(
                Q(material__name__icontains=search_query) |
                Q(damage_reason__icontains=search_query)
            ).distinct()

    # Pagination
    paginator = Paginator(damaged_qs, 20)
    page_number = request.GET.get('page')
    damaged_page = paginator.get_page(page_number)

    # Forms
    form = DamageMaterialForm(user=request.user)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            if role != 'Branch':
                messages.error(request, "Only Branch users can report Damaged Materials.")
                return redirect('damaged_materials')
            
            form = DamageMaterialForm(request.POST, user=request.user)
            if form.is_valid():
                try:
                    with transaction.atomic():
                        dm = form.save(commit=False)
                        dm.branch_user = request.user
                        dm.status = 'Pending'
                        dm.save()
                        
                        # Lock / Retire MacSerial if it exists
                        if dm.mac_serial:
                            dm.mac_serial.status = 'Retired'
                            dm.mac_serial.save()
                            
                        messages.success(request, "Damaged Material logged successfully!")
                except Exception as e:
                    messages.error(request, f"Error: {str(e)}")
                return redirect('damaged_materials')
            else:
                for field, errors in form.errors.items():
                    if errors:
                        messages.error(request, f"{field}: {errors[0]}")
                        break
                return redirect('damaged_materials')
                
        elif action == 'edit':
            if role not in ['Branch', 'Admin', 'Storekeeper', 'NOC']:
                messages.error(request, "Permission denied.")
                return redirect('damaged_materials')
                
            dm_id = request.POST.get('dm_id')
            try:
                if role == 'Branch':
                    dm = DamageMaterial.objects.get(pk=dm_id, branch_user=request.user)
                elif role == 'NOC':
                    dm = DamageMaterial.objects.get(pk=dm_id, material__category='Internet', material__created_by=request.user)
                else:
                    dm = DamageMaterial.objects.get(pk=dm_id)
                
                old_mac = dm.mac_serial
                
                # Verify owner for form context
                form = DamageMaterialForm(request.POST, instance=dm, user=dm.branch_user)
                if form.is_valid():
                    try:
                        with transaction.atomic():
                            new_dm = form.save(commit=False)
                            new_mac = form.cleaned_data.get('mac_serial')
                            
                            # If MAC serial changed, restore old one
                            if old_mac and old_mac != new_mac:
                                old_mac.status = 'Active'
                                old_mac.save()
                                
                            # Lock/Retire the new MAC serial
                            if new_mac:
                                new_mac.status = 'Retired'
                                new_mac.save()
                                
                            new_dm.save()
                            messages.success(request, "Damaged Material updated successfully!")
                    except Exception as e:
                        messages.error(request, f"Error updating record: {str(e)}")
                    return redirect('damaged_materials')
                else:
                    for field, errors in form.errors.items():
                        if errors:
                            messages.error(request, f"{field}: {errors[0]}")
                            break
                    return redirect('damaged_materials')
            except DamageMaterial.DoesNotExist:
                messages.error(request, "Record not found or access denied.")
                return redirect('damaged_materials')
                
        elif action == 'delete':
            if role != 'Branch':
                messages.error(request, "Only Branch users can delete Damaged Materials.")
                return redirect('damaged_materials')
                
            dm_id = request.POST.get('dm_id')
            try:
                dm = DamageMaterial.objects.get(pk=dm_id, branch_user=request.user)
                if dm.status == 'Pending':
                    try:
                        with transaction.atomic():
                            if dm.mac_serial:
                                dm.mac_serial.status = 'Active'
                                dm.mac_serial.save()
                            dm.delete()
                        messages.success(request, "Damaged Material record deleted.")
                    except Exception as e:
                        messages.error(request, f"Error: {str(e)}")
                else:
                    messages.error(request, "You cannot delete a record that has already been processed.")
                return redirect('damaged_materials')
            except DamageMaterial.DoesNotExist:
                messages.error(request, "Record not found or access denied.")
                return redirect('damaged_materials')

        # Admin/Storekeeper actions: confirm/reject
        elif action in ['confirm', 'reject']:
            if role not in ['Admin', 'Storekeeper', 'NOC']:
                messages.error(request, "Permission denied.")
                return redirect('damaged_materials')
                
            dm_id = request.POST.get('dm_id')
            admin_note = request.POST.get('admin_note', '').strip()
            try:
                if role == 'NOC':
                    dm = DamageMaterial.objects.get(pk=dm_id, material__category='Internet', material__created_by=request.user)
                else:
                    dm = DamageMaterial.objects.get(pk=dm_id)
                
                try:
                    with transaction.atomic():
                        dm.status = 'Confirmed' if action == 'confirm' else 'Rejected'
                        dm.admin_note = admin_note
                        dm.confirmed_by = request.user
                        dm.confirmed_at = timezone.now()
                        dm.save()
                        
                        if dm.mac_serial:
                            if action == 'confirm':
                                dm.mac_serial.status = 'Retired'
                            else:
                                dm.mac_serial.status = 'Active'
                            dm.mac_serial.save()
                            
                    messages.success(request, f"Damaged Material record has been {dm.status.lower()}!")
                except Exception as e:
                    messages.error(request, f"Error: {str(e)}")
                return redirect('damaged_materials')
            except DamageMaterial.DoesNotExist:
                messages.error(request, "Record not found.")
                return redirect('damaged_materials')

    template_name = 'noc/damaged_materials.html' if role == 'NOC' else 'inventory/damaged_materials.html'
    return render(request, template_name, {
        'damaged_materials': damaged_page,
        'form': form,
        'role': role,
        'page_obj': damaged_page,
        'branch_users': branch_users,
        'search_query': search_query,
    })

@login_required
def report_damage_auto(request):
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    if role != 'Branch':
        return JsonResponse({'success': False, 'error': 'Only Branch users can report damaged materials.'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else request.POST
    except Exception:
        data = request.POST

    material_id = data.get('material_id')
    mac_serial_id = data.get('mac_serial_id')
    quantity = int(data.get('quantity', 1))
    damage_reason = data.get('damage_reason', 'Reported directly from in-stock approved materials list.').strip()

    if not material_id:
        return JsonResponse({'success': False, 'error': 'Material ID is required.'}, status=400)

    try:
        material = Material.objects.get(id=material_id)
    except Material.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Material not found.'}, status=404)

    mac_serial_obj = None
    if mac_serial_id:
        try:
            mac_serial_obj = MacSerialNumber.objects.get(id=mac_serial_id, assigned_to=request.user)
            quantity = 1
        except MacSerialNumber.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Selected Serial Number not found or not assigned to you.'}, status=400)

    # Validate that they actually have enough in-stock quantity to mark as damaged!
    from django.db.models import Sum
    total_approved = MaterialRequest.objects.filter(
        requester=request.user,
        material=material,
        status='Received'
    ).aggregate(total=Sum('quantity'))['total'] or 0
    
    refundable_qty = RefundableMaterial.objects.filter(
        branch_user=request.user,
        material_name=material.name
    ).aggregate(total=Sum('quantity'))['total'] or 0
    
    damaged_qty = DamageMaterial.objects.filter(
        branch_user=request.user,
        material=material,
        status__in=['Pending', 'Confirmed']
    ).aggregate(total=Sum('quantity'))['total'] or 0

    used_qty = refundable_qty + damaged_qty
    available = total_approved - used_qty

    if quantity > available:
        return JsonResponse({
            'success': False, 
            'error': f'Insufficient available stock. Available: {available}, requested: {quantity}.'
        }, status=400)

    try:
        with transaction.atomic():
            # Create the DamageMaterial record!
            dm = DamageMaterial.objects.create(
                branch_user=request.user,
                material=material,
                quantity=quantity,
                damage_reason=damage_reason,
                mac_serial=mac_serial_obj,
                status='Pending'
            )
            
            # Update MacSerial status
            if mac_serial_obj:
                mac_serial_obj.status = 'Retired'
                mac_serial_obj.save()
                
            return JsonResponse({
                'success': True,
                'message': f'Successfully reported {quantity} {material.name} as damaged.'
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def get_damaged_material_api(request, pk):
    """API endpoint to get damaged material data for editing via AJAX"""
    profile = ensure_userprofile(request.user)
    role = profile.role if profile else 'Branch'

    try:
        if role in ['Admin', 'Storekeeper']:
            dm = DamageMaterial.objects.exclude(material__category='Internet').get(pk=pk)
        elif role == 'NOC':
            dm = DamageMaterial.objects.get(pk=pk, material__category='Internet', material__created_by=request.user)
        else:
            dm = DamageMaterial.objects.get(pk=pk, branch_user=request.user)
    except DamageMaterial.DoesNotExist:
        return JsonResponse({'error': 'Record not found or access denied'}, status=404)

    selection = f"s:{dm.mac_serial.id}" if dm.mac_serial else f"m:{dm.material.id}"
    
    data = {
        'id': dm.id,
        'material_selection': selection,
        'quantity': dm.quantity,
        'damage_reason': dm.damage_reason or '',
        'mac_serial': dm.mac_serial.id if dm.mac_serial else '',
        'status': dm.status,
        'admin_note': dm.admin_note or '',
    }
    return JsonResponse(data)

# Custom 404 handler
def custom_404_view(request, exception=None):
    """Render a beautiful custom 404 page."""
    context = {
        'request_path': request.path,
    }
    return render(request, '404.html', context, status=404)